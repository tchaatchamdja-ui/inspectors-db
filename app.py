import os, sqlite3, json, hashlib, secrets, csv, io, smtplib, string, random, urllib.parse
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from datetime import datetime, timedelta, timezone
from functools import wraps
from flask import Flask, request, jsonify, send_from_directory, send_file
from flask_cors import CORS
import jwt
import bcrypt
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from fpdf import FPDF

# Load .env if available
try:
    from dotenv import load_dotenv
    load_dotenv()
except ImportError:
    pass

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
_DT_FMT = "%d/%m/%Y à %H:%M"   # format date/heure — backslash hors f-string (Python 3.11)
DB_PATH = os.path.join(BASE_DIR, 'data', 'inspectors.db')
UPLOAD_DIR = os.path.join(BASE_DIR, 'uploads')
DATABASE_URL = os.environ.get('DATABASE_URL', '')
USE_POSTGRES = DATABASE_URL.startswith('postgres')
JWT_SECRET = os.environ.get('JWT_SECRET', 'uemoa-inspectors-secret-2025')
JWT_EXP_MINUTES = 20

# PostgreSQL support
if USE_POSTGRES:
    import psycopg2
    import psycopg2.extras
    # Render uses postgres:// but psycopg2 needs postgresql://
    if DATABASE_URL.startswith('postgres://'):
        DATABASE_URL = DATABASE_URL.replace('postgres://', 'postgresql://', 1)

# SMTP Configuration for email sending
SMTP_HOST = os.environ.get('SMTP_HOST', 'smtp.office365.com')
SMTP_PORT = int(os.environ.get('SMTP_PORT', 587))
SMTP_USER = os.environ.get('SMTP_USER', '')
SMTP_PASSWORD = os.environ.get('SMTP_PASSWORD', '')
SMTP_FROM = os.environ.get('SMTP_FROM', SMTP_USER)

def send_email_smtp(to_email, subject, body_html):
    """Send email via SMTP. Returns True on success, error string on failure."""
    if not SMTP_USER or not SMTP_PASSWORD:
        return "Configuration SMTP manquante (SMTP_USER/SMTP_PASSWORD)"
    try:
        msg = MIMEMultipart('alternative')
        msg['From'] = SMTP_FROM
        msg['To'] = to_email
        msg['Subject'] = subject
        msg.attach(MIMEText(body_html, 'html', 'utf-8'))
        with smtplib.SMTP(SMTP_HOST, SMTP_PORT, timeout=15) as server:
            server.starttls()
            server.login(SMTP_USER, SMTP_PASSWORD)
            server.send_message(msg)
        return True
    except Exception as e:
        return str(e)

app = Flask(__name__, static_folder='static', template_folder='templates')
CORS(app)

@app.errorhandler(404)
def not_found(e):
    if request.path.startswith('/api/'):
        return jsonify({'error': 'Route non trouvée'}), 404
    return send_from_directory('templates', 'index.html')

@app.errorhandler(500)
def server_error(e):
    return jsonify({'error': f'Erreur serveur: {str(e)}'}), 500

os.makedirs(os.path.join(BASE_DIR, 'data'), exist_ok=True)
os.makedirs(UPLOAD_DIR, exist_ok=True)

# ===== DATABASE =====
import re as _re_mod
_RE_INSERT_IGNORE = _re_mod.compile(r'INSERT\s+OR\s+IGNORE\s+INTO', _re_mod.IGNORECASE)
_RE_INSERT_REPLACE = _re_mod.compile(r'INSERT\s+OR\s+REPLACE\s+INTO', _re_mod.IGNORECASE)
_RE_GROUP_CONCAT_DISTINCT = _re_mod.compile(r'GROUP_CONCAT\(DISTINCT\s+((?:[^()]*|\([^()]*\))*)\)')
_RE_GROUP_CONCAT = _re_mod.compile(r'GROUP_CONCAT\(((?:[^()]*|\([^()]*\))*)\)')
_SQL_CONVERT_CACHE = {}

def _q(sql):
    """Convert SQLite SQL to PostgreSQL (mémoïsé)"""
    if not USE_POSTGRES:
        return sql
    cached = _SQL_CONVERT_CACHE.get(sql)
    if cached is not None:
        return cached
    out = sql.replace('?', '%s').replace("datetime('now')", 'CURRENT_TIMESTAMP')
    out = _RE_INSERT_IGNORE.sub('INSERT INTO', out)
    out = _RE_INSERT_REPLACE.sub('INSERT INTO', out)
    out = _RE_GROUP_CONCAT_DISTINCT.sub(r"STRING_AGG(DISTINCT (\1)::text, ', ')", out)
    out = _RE_GROUP_CONCAT.sub(r"STRING_AGG((\1)::text, ', ')", out)
    if len(_SQL_CONVERT_CACHE) < 2000:
        _SQL_CONVERT_CACHE[sql] = out
    return out

class PgRowWrapper(dict):
    """Wrap psycopg2 RealDictRow to support both dict-style and index access"""
    def __init__(self, row):
        super().__init__(row)

class PgCursorWrapper:
    """Wrap psycopg2 cursor to match sqlite3 API"""
    def __init__(self, conn):
        self._conn = conn
        self._cursor = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
        self._last_insert_id = None

    def execute(self, sql, params=None):
        is_ignore = bool(_RE_INSERT_IGNORE.search(sql))
        converted = _q(sql)
        if converted.strip().upper().startswith('INSERT') and 'RETURNING' not in converted.upper():
            suffix = ' ON CONFLICT DO NOTHING RETURNING id' if is_ignore else ' RETURNING id'
            self._cursor.execute(converted + suffix, params or ())
            row = self._cursor.fetchone()
            self._last_insert_id = row['id'] if row else None
        else:
            self._cursor.execute(converted, params or ())
            self._last_insert_id = None
        return self

    def executescript(self, sql):
        """Execute multiple SQL statements separated by semicolons."""
        statements = [s.strip() for s in sql.split(';') if s.strip()]
        for stmt in statements:
            self._cursor.execute(stmt)
        return self

    def fetchone(self):
        row = self._cursor.fetchone()
        return PgRowWrapper(row) if row else None

    def fetchall(self):
        return [PgRowWrapper(r) for r in self._cursor.fetchall()]

    def commit(self):
        self._conn.commit()

    def rollback(self):
        self._conn.rollback()

    def close(self):
        self._cursor.close()
        self._conn.close()

    @property
    def lastrowid(self):
        return self._last_insert_id

def get_db():
    if USE_POSTGRES:
        conn = psycopg2.connect(DATABASE_URL, sslmode='require', connect_timeout=10)
        return PgCursorWrapper(conn)
    else:
        db = sqlite3.connect(DB_PATH, timeout=10)
        db.row_factory = sqlite3.Row
        db.execute("PRAGMA foreign_keys = ON")
        db.execute("PRAGMA journal_mode = WAL")
        return db

def init_db():
    db = get_db()
    if USE_POSTGRES:
        db.executescript("""
            CREATE TABLE IF NOT EXISTS users (
                id SERIAL PRIMARY KEY,
                username TEXT UNIQUE NOT NULL,
                password TEXT NOT NULL,
                role TEXT NOT NULL DEFAULT 'National 2',
                is_active INTEGER NOT NULL DEFAULT 1,
                must_change_password INTEGER NOT NULL DEFAULT 1,
                inspector_id INTEGER,
                user_nom TEXT,
                user_etat TEXT,
                user_prenom TEXT,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            );
            CREATE TABLE IF NOT EXISTS inspectors (
                id SERIAL PRIMARY KEY,
                reference TEXT UNIQUE NOT NULL,
                nom TEXT NOT NULL,
                prenom TEXT NOT NULL,
                etat TEXT NOT NULL,
                email TEXT,
                telephone TEXT,
                cv_path TEXT,
                is_active INTEGER NOT NULL DEFAULT 1,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            );
            CREATE TABLE IF NOT EXISTS qualifications (
                id SERIAL PRIMARY KEY,
                inspector_id INTEGER NOT NULL,
                domaine TEXT NOT NULL,
                specialite TEXT NOT NULL,
                niveau TEXT NOT NULL,
                experience TEXT,
                titularisation TEXT,
                FOREIGN KEY (inspector_id) REFERENCES inspectors(id) ON DELETE CASCADE
            );
            CREATE TABLE IF NOT EXISTS activity_log (
                id SERIAL PRIMARY KEY,
                user_id INTEGER,
                action TEXT NOT NULL,
                details TEXT,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            );
            CREATE TABLE IF NOT EXISTS password_reset_requests (
                id SERIAL PRIMARY KEY,
                user_id INTEGER,
                email TEXT NOT NULL,
                phone TEXT NOT NULL DEFAULT '',
                token TEXT,
                status TEXT DEFAULT 'pending',
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            );
            CREATE TABLE IF NOT EXISTS settings (
                id SERIAL PRIMARY KEY,
                category TEXT NOT NULL,
                value TEXT NOT NULL,
                label TEXT,
                is_active INTEGER NOT NULL DEFAULT 1,
                sort_order INTEGER DEFAULT 0,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            );
            CREATE TABLE IF NOT EXISTS formateurs (
                id SERIAL PRIMARY KEY,
                reference TEXT UNIQUE NOT NULL,
                nom TEXT NOT NULL,
                prenom TEXT NOT NULL,
                etat TEXT NOT NULL,
                email TEXT,
                telephone TEXT,
                cv_path TEXT,
                is_inspecteur INTEGER NOT NULL DEFAULT 0,
                inspector_id INTEGER REFERENCES inspectors(id) ON DELETE SET NULL,
                is_active INTEGER NOT NULL DEFAULT 1,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            );
            CREATE TABLE IF NOT EXISTS formateur_competences (
                id SERIAL PRIMARY KEY,
                formateur_id INTEGER NOT NULL,
                type_competence TEXT NOT NULL,
                domaine TEXT,
                FOREIGN KEY (formateur_id) REFERENCES formateurs(id) ON DELETE CASCADE
            );
            CREATE TABLE IF NOT EXISTS formateur_formations (
                id SERIAL PRIMARY KEY,
                formateur_id INTEGER NOT NULL,
                type TEXT NOT NULL,
                description TEXT NOT NULL,
                FOREIGN KEY (formateur_id) REFERENCES formateurs(id) ON DELETE CASCADE
            );
            -- Migrations idempotentes pour bases existantes (PG 9.6+)
            ALTER TABLE users          ADD COLUMN IF NOT EXISTS user_nom TEXT;
            ALTER TABLE users          ADD COLUMN IF NOT EXISTS user_prenom TEXT;
            ALTER TABLE users          ADD COLUMN IF NOT EXISTS user_etat TEXT;
            ALTER TABLE qualifications ADD COLUMN IF NOT EXISTS titularisation TEXT;
            ALTER TABLE formateurs     ADD COLUMN IF NOT EXISTS inspector_id INTEGER;
            CREATE INDEX IF NOT EXISTS idx_qualifications_inspector ON qualifications(inspector_id);
            CREATE INDEX IF NOT EXISTS idx_qualifications_domaine ON qualifications(domaine);
            CREATE INDEX IF NOT EXISTS idx_qualifications_niveau ON qualifications(niveau);
            CREATE INDEX IF NOT EXISTS idx_users_inspector ON users(inspector_id);
            CREATE INDEX IF NOT EXISTS idx_inspectors_active_etat ON inspectors(is_active, etat);
            CREATE INDEX IF NOT EXISTS idx_inspectors_nom ON inspectors(nom);
            CREATE INDEX IF NOT EXISTS idx_formateur_competences_formateur ON formateur_competences(formateur_id);
            CREATE INDEX IF NOT EXISTS idx_formateur_competences_type ON formateur_competences(type_competence);
            CREATE INDEX IF NOT EXISTS idx_formateur_formations_formateur ON formateur_formations(formateur_id);
            CREATE INDEX IF NOT EXISTS idx_formateurs_active_etat ON formateurs(is_active, etat);
            CREATE INDEX IF NOT EXISTS idx_formateurs_nom ON formateurs(nom);
            CREATE INDEX IF NOT EXISTS idx_formateurs_inspector ON formateurs(inspector_id);
            CREATE INDEX IF NOT EXISTS idx_settings_category_active ON settings(category, is_active);
            CREATE INDEX IF NOT EXISTS idx_activity_log_user ON activity_log(user_id, created_at);
        """)
        db.commit()
    else:
        db.executescript("""
            CREATE TABLE IF NOT EXISTS users (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                username TEXT UNIQUE NOT NULL,
                password TEXT NOT NULL,
                role TEXT NOT NULL DEFAULT 'National 2',
                is_active INTEGER NOT NULL DEFAULT 1,
                must_change_password INTEGER NOT NULL DEFAULT 1,
                inspector_id INTEGER,
                user_nom TEXT,
                user_etat TEXT,
                created_at TEXT DEFAULT (datetime('now')),
                updated_at TEXT DEFAULT (datetime('now'))
            );
            CREATE TABLE IF NOT EXISTS inspectors (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                reference TEXT UNIQUE NOT NULL,
                nom TEXT NOT NULL,
                prenom TEXT NOT NULL,
                etat TEXT NOT NULL,
                email TEXT,
                telephone TEXT,
                cv_path TEXT,
                is_active INTEGER NOT NULL DEFAULT 1,
                created_at TEXT DEFAULT (datetime('now')),
                updated_at TEXT DEFAULT (datetime('now'))
            );
            CREATE TABLE IF NOT EXISTS qualifications (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                inspector_id INTEGER NOT NULL,
                domaine TEXT NOT NULL,
                specialite TEXT NOT NULL,
                niveau TEXT NOT NULL,
                experience TEXT,
                titularisation TEXT,
                FOREIGN KEY (inspector_id) REFERENCES inspectors(id) ON DELETE CASCADE
            );
            CREATE TABLE IF NOT EXISTS activity_log (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER,
                action TEXT NOT NULL,
                details TEXT,
                created_at TEXT DEFAULT (datetime('now'))
            );
            CREATE TABLE IF NOT EXISTS password_reset_requests (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER,
                email TEXT NOT NULL,
                phone TEXT NOT NULL,
                token TEXT,
                status TEXT DEFAULT 'pending',
                created_at TEXT DEFAULT (datetime('now'))
            );
            CREATE TABLE IF NOT EXISTS settings (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                category TEXT NOT NULL,
                value TEXT NOT NULL,
                label TEXT,
                is_active INTEGER NOT NULL DEFAULT 1,
                sort_order INTEGER DEFAULT 0,
                created_at TEXT DEFAULT (datetime('now')),
                updated_at TEXT DEFAULT (datetime('now'))
            );
            CREATE TABLE IF NOT EXISTS formateurs (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                reference TEXT UNIQUE NOT NULL,
                nom TEXT NOT NULL,
                prenom TEXT NOT NULL,
                etat TEXT NOT NULL,
                email TEXT,
                telephone TEXT,
                cv_path TEXT,
                is_inspecteur INTEGER NOT NULL DEFAULT 0,
                inspector_id INTEGER REFERENCES inspectors(id) ON DELETE SET NULL,
                is_active INTEGER NOT NULL DEFAULT 1,
                created_at TEXT DEFAULT (datetime('now')),
                updated_at TEXT DEFAULT (datetime('now'))
            );
            CREATE TABLE IF NOT EXISTS formateur_competences (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                formateur_id INTEGER NOT NULL,
                type_competence TEXT NOT NULL,
                domaine TEXT,
                FOREIGN KEY (formateur_id) REFERENCES formateurs(id) ON DELETE CASCADE
            );
            CREATE TABLE IF NOT EXISTS formateur_formations (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                formateur_id INTEGER NOT NULL,
                type TEXT NOT NULL,
                description TEXT NOT NULL,
                FOREIGN KEY (formateur_id) REFERENCES formateurs(id) ON DELETE CASCADE
            );
        """)
        # Migrate: add columns if missing
        cols = [c[1] for c in db.execute("PRAGMA table_info(users)").fetchall()]
        if 'user_nom' not in cols:
            db.execute("ALTER TABLE users ADD COLUMN user_nom TEXT")
        if 'user_etat' not in cols:
            db.execute("ALTER TABLE users ADD COLUMN user_etat TEXT")
        if 'user_prenom' not in cols:
            db.execute("ALTER TABLE users ADD COLUMN user_prenom TEXT")
        qual_cols = [c[1] for c in db.execute("PRAGMA table_info(qualifications)").fetchall()]
        if 'titularisation' not in qual_cols:
            db.execute("ALTER TABLE qualifications ADD COLUMN titularisation TEXT")
        # Lien inspecteur <-> formateur (Option 2)
        frm_cols = [c[1] for c in db.execute("PRAGMA table_info(formateurs)").fetchall()]
        if 'inspector_id' not in frm_cols:
            db.execute("ALTER TABLE formateurs ADD COLUMN inspector_id INTEGER")
        # Index pour performance
        db.executescript("""
            CREATE INDEX IF NOT EXISTS idx_qualifications_inspector ON qualifications(inspector_id);
            CREATE INDEX IF NOT EXISTS idx_qualifications_domaine ON qualifications(domaine);
            CREATE INDEX IF NOT EXISTS idx_qualifications_niveau ON qualifications(niveau);
            CREATE INDEX IF NOT EXISTS idx_users_inspector ON users(inspector_id);
            CREATE INDEX IF NOT EXISTS idx_inspectors_active_etat ON inspectors(is_active, etat);
            CREATE INDEX IF NOT EXISTS idx_inspectors_nom ON inspectors(nom);
            CREATE INDEX IF NOT EXISTS idx_formateur_competences_formateur ON formateur_competences(formateur_id);
            CREATE INDEX IF NOT EXISTS idx_formateur_competences_type ON formateur_competences(type_competence);
            CREATE INDEX IF NOT EXISTS idx_formateur_formations_formateur ON formateur_formations(formateur_id);
            CREATE INDEX IF NOT EXISTS idx_formateurs_active_etat ON formateurs(is_active, etat);
            CREATE INDEX IF NOT EXISTS idx_formateurs_nom ON formateurs(nom);
            CREATE INDEX IF NOT EXISTS idx_formateurs_inspector ON formateurs(inspector_id);
            CREATE INDEX IF NOT EXISTS idx_settings_category_active ON settings(category, is_active);
            CREATE INDEX IF NOT EXISTS idx_activity_log_user ON activity_log(user_id, created_at);
        """)
        db.commit()
    # Migration auto-lien inspecteur <-> formateur (idempotent)
    try:
        # Lien par nom+prenom+etat
        db.execute("""
            UPDATE formateurs SET inspector_id = (
              SELECT i.id FROM inspectors i
              WHERE UPPER(TRIM(i.nom))    = UPPER(TRIM(formateurs.nom))
                AND UPPER(TRIM(i.prenom)) = UPPER(TRIM(formateurs.prenom))
                AND TRIM(i.etat)          = TRIM(formateurs.etat)
                AND i.is_active = 1
              LIMIT 1
            )
            WHERE formateurs.is_active = 1 AND formateurs.inspector_id IS NULL
        """)
        # Lien par email (fallback pour cas avec prénoms variant)
        db.execute("""
            UPDATE formateurs SET inspector_id = (
              SELECT i.id FROM inspectors i
              WHERE LOWER(TRIM(i.email)) = LOWER(TRIM(formateurs.email))
                AND i.email IS NOT NULL AND i.email != ''
                AND i.is_active = 1
              LIMIT 1
            )
            WHERE formateurs.is_active = 1 AND formateurs.inspector_id IS NULL
              AND formateurs.email IS NOT NULL AND formateurs.email != ''
        """)
        db.execute("UPDATE formateurs SET is_inspecteur = 1 WHERE inspector_id IS NOT NULL AND is_inspecteur = 0")
        db.commit()
    except Exception as _e:
        print(f"[migration link] avertissement : {_e}")
    # Create admin
    admin = db.execute("SELECT id FROM users WHERE username = 'Admin'").fetchone()
    if not admin:
        pw = bcrypt.hashpw('admin'.encode(), bcrypt.gensalt()).decode()
        db.execute("INSERT INTO users (username, password, role, must_change_password) VALUES (?, ?, 'Administrateur', 0)", ('Admin', pw))
    # Seed default settings if empty
    existing = db.execute("SELECT COUNT(*) as c FROM settings").fetchone()
    if existing['c'] == 0:
        defaults = [
            ('etat', 'Burkina Faso'), ('etat', 'B\u00e9nin'), ('etat', "C\u00f4te d'Ivoire"),
            ('etat', 'Guin\u00e9e-Bissau'), ('etat', 'Mali'), ('etat', 'Mauritanie'),
            ('etat', 'Niger'), ('etat', 'S\u00e9n\u00e9gal'), ('etat', 'Togo'), ('etat', 'UEMOA'),
            ('domaine', 'PEL', 'D\u00e9livrance des licences et formation du personnel'),
            ('domaine', 'OPS', 'Exploitation technique des a\u00e9ronefs'),
            ('domaine', 'AIR', 'Navigabilit\u00e9 des a\u00e9ronefs'),
            ('domaine', 'AIG', "Enqu\u00eates sur les accidents et incidents d'aviation"),
            ('domaine', 'AGA', 'A\u00e9rodromes et aides au sol'),
            ('domaine', 'PNS', 'Gestion de la s\u00e9curit\u00e9'),
            ('domaine', 'AVSEC', 'Suret\u00e9'),
            ('formateur', 'Instructeur'), ('formateur', 'D\u00e9veloppeur de Cours'),
            ('niveau', 'Inspecteur Stagiaire'), ('niveau', 'Inspecteur Titulaire'),
            ('niveau', 'Inspecteur Principal'), ('niveau', 'Inspecteur Senior'),
            ('niveau', 'Enqu\u00eateur Technique'), ('niveau', 'Enqu\u00eateur de Premi\u00e8re Information'),
            ('niveau', 'Enqu\u00eateur Confirm\u00e9'),
        ]
        for i, item in enumerate(defaults):
            cat = item[0]
            val = item[1]
            lbl = item[2] if len(item) > 2 else None
            db.execute("INSERT INTO settings (category, value, label, sort_order) VALUES (?, ?, ?, ?)",
                       (cat, val, lbl, i))
    db.commit()
    db.close()

def log_activity(db, user_id, action, details=''):
    db.execute("INSERT INTO activity_log (user_id, action, details) VALUES (?, ?, ?)", (user_id, action, details))

def generate_reference(db):
    row = db.execute("SELECT reference FROM inspectors ORDER BY id DESC LIMIT 1").fetchone()
    num = 1
    if row:
        try: num = int(row['reference'].split('-')[-1]) + 1
        except: pass
    return f"UEMOA-INS-{num:04d}"

# ===== AUTH =====
def create_token(user_dict):
    payload = {
        'id': user_dict['id'], 'username': user_dict['username'],
        'role': user_dict['role'], 'inspector_id': user_dict.get('inspector_id'),
        'exp': datetime.now(timezone.utc) + timedelta(minutes=JWT_EXP_MINUTES)
    }
    return jwt.encode(payload, JWT_SECRET, algorithm='HS256')

def auth_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        token = request.headers.get('Authorization', '').replace('Bearer ', '')
        if not token:
            return jsonify({'error': 'Non autorisé'}), 401
        try:
            data = jwt.decode(token, JWT_SECRET, algorithms=['HS256'])
            request.user = data
        except jwt.ExpiredSignatureError:
            return jsonify({'error': 'Session expirée, veuillez vous reconnecter'}), 401
        except:
            return jsonify({'error': 'Token invalide'}), 401
        return f(*args, **kwargs)
    return decorated

def require_role(*roles):
    def decorator(f):
        @wraps(f)
        def decorated(*args, **kwargs):
            if request.user.get('role') not in roles:
                return jsonify({'error': 'Accès non autorisé'}), 403
            return f(*args, **kwargs)
        return decorated
    return decorator

# ===== STATIC FILES =====
@app.route('/')
def index():
    return send_from_directory('templates', 'index.html')

@app.route('/uploads/<path:filename>')
def uploaded_file(filename):
    return send_from_directory(UPLOAD_DIR, filename)

# ===== AUTH ROUTES =====
@app.route('/api/auth/login', methods=['POST'])
def login():
    data = request.json
    username = data.get('username', '')
    password = data.get('password', '')
    if not username or not password:
        return jsonify({'error': 'Identifiants requis'}), 400

    db = get_db()
    user = db.execute("SELECT * FROM users WHERE username = ? AND is_active = 1", (username,)).fetchone()
    if not user:
        db.close()
        return jsonify({'error': 'Identifiants incorrects'}), 401

    if not bcrypt.checkpw(password.encode(), user['password'].encode()):
        db.close()
        return jsonify({'error': 'Identifiants incorrects'}), 401

    log_activity(db, user['id'], 'LOGIN', f"Connexion de {username}")
    db.commit()

    user_dict = dict(user)
    token = create_token(user_dict)
    db.close()
    return jsonify({
        'token': token,
        'user': {
            'id': user_dict['id'], 'username': user_dict['username'],
            'role': user_dict['role'], 'mustChangePassword': user_dict['must_change_password'] == 1,
            'inspectorId': user_dict['inspector_id']
        }
    })

@app.route('/api/auth/change-password', methods=['POST'])
@auth_required
def change_password():
    data = request.json
    new_pw = data.get('newPassword', '')
    if len(new_pw) < 6:
        return jsonify({'error': 'Le mot de passe doit contenir au moins 6 caractères'}), 400

    db = get_db()
    user = db.execute("SELECT * FROM users WHERE id = ?", (request.user['id'],)).fetchone()
    if not user:
        db.close()
        return jsonify({'error': 'Utilisateur non trouvé'}), 404

    if not user['must_change_password'] and data.get('currentPassword'):
        if not bcrypt.checkpw(data['currentPassword'].encode(), user['password'].encode()):
            db.close()
            return jsonify({'error': 'Mot de passe actuel incorrect'}), 401

    hashed = bcrypt.hashpw(new_pw.encode(), bcrypt.gensalt()).decode()
    db.execute("UPDATE users SET password = ?, must_change_password = 0, updated_at = datetime('now') WHERE id = ?", (hashed, request.user['id']))
    log_activity(db, request.user['id'], 'PASSWORD_CHANGE', 'Mot de passe modifié')
    db.commit()

    user_dict = dict(user)
    user_dict['must_change_password'] = 0
    token = create_token(user_dict)
    db.close()
    return jsonify({'message': 'Mot de passe modifié avec succès', 'token': token})

@app.route('/api/auth/request-reset', methods=['POST'])
def request_reset():
    data = request.json
    email = data.get('email', '')
    if not email:
        return jsonify({'error': 'Email requis'}), 400

    db = get_db()
    user = db.execute("SELECT * FROM users WHERE username = ?", (email,)).fetchone()

    recent = db.execute("SELECT COUNT(*) as cnt FROM password_reset_requests WHERE email = ? AND created_at > datetime('now', '-1 hour')", (email,)).fetchone()

    if not user:
        db.execute("INSERT INTO password_reset_requests (user_id, email, phone, status) VALUES (NULL, ?, '', 'failed')", (email,))
        db.commit()
        if recent['cnt'] >= 2:
            admin = db.execute("SELECT id FROM users WHERE role = 'Administrateur'").fetchone()
            if admin:
                log_activity(db, admin['id'], 'RESET_ALERT', f"Tentatives échouées de réinitialisation pour: {email}")
                db.commit()
            db.close()
            return jsonify({'error': "Informations incorrectes. Veuillez contacter l'administrateur.", 'contactAdmin': True}), 400
        db.close()
        return jsonify({'error': 'Informations incorrectes'}), 400

    token = secrets.token_urlsafe(20)
    db.execute("INSERT INTO password_reset_requests (user_id, email, phone, token, status) VALUES (?, ?, '', ?, 'sent')", (user['id'], email, token))
    log_activity(db, user['id'], 'RESET_REQUEST', 'Demande de réinitialisation envoyée')
    admin = db.execute("SELECT id FROM users WHERE role = 'Administrateur'").fetchone()
    if admin:
        log_activity(db, admin['id'], 'RESET_NOTIFICATION', f"Demande de réinitialisation de: {email}")
    db.commit()
    db.close()
    return jsonify({'message': "Votre demande a été envoyée. L'administrateur a été notifié et procédera à la réinitialisation de votre mot de passe."})

@app.route('/api/auth/me', methods=['GET'])
@auth_required
def get_me():
    db = get_db()
    user = db.execute("SELECT id, username, role, must_change_password, inspector_id, is_active FROM users WHERE id = ?", (request.user['id'],)).fetchone()
    db.close()
    if not user: return jsonify({'error': 'Non trouvé'}), 404
    return jsonify(dict(user))

# ===== INSPECTOR ROUTES =====
@app.route('/api/inspectors/stats', methods=['GET'])
@auth_required
def inspector_stats():
    db = get_db()
    # Build filter conditions matching list_inspectors logic
    where = []
    params = []
    status = request.args.get('status', 'active')
    if status == 'active':
        where.append('i.is_active = 1')
    elif status == 'inactive':
        where.append('i.is_active = 0')
    etat = request.args.get('etat', '')
    if etat:
        where.append('i.etat = ?')
        params.append(etat)
    search = request.args.get('search', '')
    if search:
        where.append('(i.nom LIKE ? OR i.prenom LIKE ?)')
        params.extend([f'%{search}%', f'%{search}%'])
    domaine = request.args.get('domaine', '')
    niveau = request.args.get('niveau', '')
    experience = request.args.get('experience', '')
    join_qual = ''
    qual_where = []
    qual_params = []
    if domaine or niveau or experience:
        join_qual = 'JOIN qualifications q ON q.inspector_id = i.id'
        if domaine:
            qual_where.append('q.domaine = ?')
            qual_params.append(domaine)
        if niveau:
            qual_where.append('q.niveau = ?')
            qual_params.append(niveau)
        if experience:
            exp_map = {
                'less1': ["Moins d'un an", "0,5 an", "04 mois", "09 mois"],
                '1to3': ["01 an", "01 ans", "02 ans"],
                '3to5': ["03 ans", "04 ans"],
                '5to8': ["05 ans", "06 ans", "07 ans", "7 ans"],
                '8to10': ["08 ans", "09 ans"],
                '10to15': ["10 ans", "11 ans", "12 ans", "13 ans", "14 ans", "plus de 10 ans"],
                '15plus': ["17 ans"]
            }
            vals = exp_map.get(experience, [])
            if vals:
                qual_where.append(f"q.experience IN ({','.join(['?'] * len(vals))})")
                qual_params.extend(vals)

    if not where:
        where = ['1=1']
    all_where = ' AND '.join(where + qual_where)
    all_params = params + qual_params

    total = db.execute(f"SELECT COUNT(DISTINCT i.id) as count FROM inspectors i {join_qual} WHERE {all_where}", all_params).fetchone()['count']
    by_state = [dict(r) for r in db.execute(f"SELECT i.etat, COUNT(DISTINCT i.id) as count FROM inspectors i {join_qual} WHERE {all_where} GROUP BY i.etat ORDER BY i.etat", all_params).fetchall()]
    by_domain = [dict(r) for r in db.execute(f"SELECT q2.domaine, COUNT(DISTINCT q2.inspector_id) as count FROM qualifications q2 JOIN inspectors i2 ON q2.inspector_id = i2.id WHERE i2.id IN (SELECT DISTINCT i.id FROM inspectors i {join_qual} WHERE {all_where}) GROUP BY q2.domaine", all_params).fetchall()]
    by_level = [dict(r) for r in db.execute(f"SELECT q2.niveau, COUNT(DISTINCT q2.inspector_id) as count FROM qualifications q2 JOIN inspectors i2 ON q2.inspector_id = i2.id WHERE i2.id IN (SELECT DISTINCT i.id FROM inspectors i {join_qual} WHERE {all_where}) GROUP BY q2.niveau ORDER BY count DESC", all_params).fetchall()]
    db.close()
    return jsonify({'total': total, 'byState': by_state, 'byDomain': by_domain, 'byLevel': by_level})

@app.route('/api/inspectors', methods=['GET'])
@auth_required
def list_inspectors():
    etat = request.args.get('etat', '')
    domaine = request.args.get('domaine', '')
    niveau = request.args.get('niveau', '')
    experience = request.args.get('experience', '')
    search = request.args.get('search', '')
    status = request.args.get('status', 'active')
    page = int(request.args.get('page', 1))
    limit = int(request.args.get('limit', 50))

    where = []
    params = []

    if status == 'active':
        where.append('i.is_active = 1')
    elif status == 'inactive':
        where.append('i.is_active = 0')
    # status == 'all' -> no filter

    if not where:
        where = ['1=1']

    if etat:
        where.append('i.etat = ?')
        params.append(etat)
    if search:
        where.append('(i.nom LIKE ? OR i.prenom LIKE ?)')
        params.extend([f'%{search}%', f'%{search}%'])

    join_qual = ''
    if domaine or niveau or experience:
        join_qual = 'JOIN qualifications q ON q.inspector_id = i.id'
        if domaine:
            where.append('q.domaine = ?')
            params.append(domaine)
        if niveau:
            where.append('q.niveau = ?')
            params.append(niveau)
        if experience:
            exp_map = {
                'less1': ["Moins d'un an", "0,5 an", "04 mois", "09 mois"],
                '1to3': ["01 an", "01 ans", "02 ans"],
                '3to5': ["03 ans", "04 ans"],
                '5to8': ["05 ans", "06 ans", "07 ans", "7 ans"],
                '8to10': ["08 ans", "09 ans"],
                '10to15': ["10 ans", "11 ans", "12 ans", "13 ans", "14 ans", "plus de 10 ans"],
                '15plus': ["17 ans"]
            }
            vals = exp_map.get(experience, [])
            if vals:
                placeholders = ','.join(['?'] * len(vals))
                where.append(f'q.experience IN ({placeholders})')
                params.extend(vals)

    where_clause = ' AND '.join(where)
    offset = (page - 1) * limit

    db = get_db()
    count_sql = f"SELECT COUNT(DISTINCT i.id) as total FROM inspectors i {join_qual} WHERE {where_clause}"
    total = db.execute(count_sql, params).fetchone()['total']

    sql = f"SELECT DISTINCT i.* FROM inspectors i {join_qual} WHERE {where_clause} ORDER BY i.etat, i.nom LIMIT ? OFFSET ?"
    inspectors = [dict(r) for r in db.execute(sql, params + [limit, offset]).fetchall()]

    # Batch fetch qualifications + users (évite N+1 — 2 requêtes au total au lieu de 2*N)
    ids = [ins['id'] for ins in inspectors]
    quals_by = {}
    user_set = set()
    if ids:
        placeholders = ','.join(['?'] * len(ids))
        for q in db.execute(f"SELECT * FROM qualifications WHERE inspector_id IN ({placeholders})", ids).fetchall():
            quals_by.setdefault(q['inspector_id'], []).append(dict(q))
        for u in db.execute(f"SELECT inspector_id FROM users WHERE inspector_id IN ({placeholders})", ids).fetchall():
            user_set.add(u['inspector_id'])
    for ins in inspectors:
        ins['qualifications'] = quals_by.get(ins['id'], [])
        ins['has_user'] = ins['id'] in user_set

    db.close()
    return jsonify({'inspectors': inspectors, 'total': total, 'page': page, 'totalPages': (total + limit - 1) // limit})

@app.route('/api/inspectors/<int:id>', methods=['GET'])
@auth_required
def get_inspector(id):
    db = get_db()
    ins = db.execute("SELECT * FROM inspectors WHERE id = ?", (id,)).fetchone()
    if not ins:
        db.close()
        return jsonify({'error': 'Non trouvé'}), 404
    result = dict(ins)
    result['qualifications'] = [dict(q) for q in db.execute("SELECT * FROM qualifications WHERE inspector_id = ?", (id,)).fetchall()]
    # Lien formateur (Option 2)
    frm = db.execute("SELECT id, reference, nom, prenom, etat, email FROM formateurs WHERE inspector_id = ? AND is_active = 1", (id,)).fetchone()
    result['formateur'] = dict(frm) if frm else None
    result['is_formateur'] = bool(frm)
    db.close()
    return jsonify(result)

# ---- Helpers pour lien inspecteur <-> formateur ----
def _find_formateur_match(db, nom, prenom, etat, email=None, exclude_inspector_id=None):
    """Trouve un formateur potentiellement correspondant (non encore lié à un autre inspecteur)."""
    rows = db.execute(
        "SELECT id, reference, nom, prenom, etat, email, inspector_id FROM formateurs"
        " WHERE is_active = 1"
        " AND (inspector_id IS NULL OR inspector_id = ?)"
        " AND UPPER(TRIM(nom)) = UPPER(TRIM(?))"
        " AND UPPER(TRIM(prenom)) = UPPER(TRIM(?))"
        " AND TRIM(etat) = TRIM(?)"
        " LIMIT 1",
        (exclude_inspector_id or -1, nom, prenom, etat)
    ).fetchone()
    if rows:
        return dict(rows)
    if email:
        rows = db.execute(
            "SELECT id, reference, nom, prenom, etat, email, inspector_id FROM formateurs"
            " WHERE is_active = 1 AND (inspector_id IS NULL OR inspector_id = ?)"
            " AND LOWER(TRIM(email)) = LOWER(TRIM(?)) AND email IS NOT NULL AND email != ''"
            " LIMIT 1",
            (exclude_inspector_id or -1, email)
        ).fetchone()
        if rows:
            return dict(rows)
    return None

def _handle_formateur_link(db, inspector_id, nom, prenom, etat, email, telephone):
    """Crée/lie/délie un formateur en fonction des champs du formulaire.
    Retourne (formateur_id, action) où action ∈ {'linked','created','unlinked','none'}."""
    is_formateur = request.form.get('is_formateur', '').strip().lower() in ('true', '1', 'on', 'yes')
    existing = db.execute("SELECT id FROM formateurs WHERE inspector_id = ?", (inspector_id,)).fetchone()
    if not is_formateur:
        if existing:
            db.execute("UPDATE formateurs SET inspector_id = NULL, is_inspecteur = 0, updated_at = datetime('now') WHERE id = ?", (existing['id'],))
            return (None, 'unlinked')
        return (None, 'none')

    # is_formateur = True
    if existing:
        # Déjà lié — synchroniser les champs identifiants
        db.execute("UPDATE formateurs SET nom = ?, prenom = ?, etat = ?, email = ?, telephone = ?, is_inspecteur = 1, updated_at = datetime('now') WHERE id = ?",
                   (nom, prenom, etat, email, telephone, existing['id']))
        return (existing['id'], 'none')

    explicit_fid = request.form.get('formateur_id', '').strip()
    if explicit_fid:
        try:
            fid = int(explicit_fid)
            chk = db.execute("SELECT id, inspector_id FROM formateurs WHERE id = ?", (fid,)).fetchone()
            if chk and (not chk['inspector_id'] or chk['inspector_id'] == inspector_id):
                db.execute("UPDATE formateurs SET inspector_id = ?, is_inspecteur = 1, nom = ?, prenom = ?, etat = ?, email = ?, telephone = ?, updated_at = datetime('now') WHERE id = ?",
                           (inspector_id, nom, prenom, etat, email, telephone, fid))
                return (fid, 'linked')
        except (ValueError, TypeError):
            pass

    # Auto-match par nom+prenom+etat
    match = _find_formateur_match(db, nom, prenom, etat, email)
    if match:
        db.execute("UPDATE formateurs SET inspector_id = ?, is_inspecteur = 1, email = ?, telephone = ?, updated_at = datetime('now') WHERE id = ?",
                   (inspector_id, email, telephone, match['id']))
        return (match['id'], 'linked')

    # Création d'un nouveau formateur
    ref = generate_formateur_reference(db)
    cursor = db.execute("INSERT INTO formateurs (reference, nom, prenom, etat, email, telephone, is_inspecteur, inspector_id) VALUES (?, ?, ?, ?, ?, ?, 1, ?)",
                        (ref, nom, prenom, etat, email or None, telephone or None, inspector_id))
    fid = cursor.lastrowid
    if not fid:
        fid = db.execute("SELECT id FROM formateurs WHERE reference = ?", (ref,)).fetchone()['id']
    # Compétences
    comps_json = request.form.get('formateur_competences', '[]')
    try:
        for c in json.loads(comps_json):
            t = (c.get('type_competence') or '').strip()
            d = (c.get('domaine') or '').strip() or None
            if t:
                db.execute("INSERT INTO formateur_competences (formateur_id, type_competence, domaine) VALUES (?, ?, ?)", (fid, t, d))
    except Exception:
        pass
    # Formations délivrées / développées
    for kind, key in (('delivree', 'formateur_formations_delivrees'), ('developpee', 'formateur_formations_developpees')):
        try:
            for desc in json.loads(request.form.get(key, '[]')):
                desc = (desc or '').strip()
                if desc:
                    db.execute("INSERT INTO formateur_formations (formateur_id, type, description) VALUES (?, ?, ?)", (fid, kind, desc))
        except Exception:
            pass
    return (fid, 'created')

@app.route('/api/formateurs/match', methods=['GET'])
@auth_required
def formateur_match():
    """Renvoie un formateur potentiellement correspondant à un inspecteur (par nom+prenom+etat ou email)."""
    if request.user['role'] not in ('National 1', 'Régional', 'Administrateur'):
        return jsonify({'error': 'Non autorisé'}), 403
    nom = (request.args.get('nom') or '').strip()
    prenom = (request.args.get('prenom') or '').strip()
    etat = (request.args.get('etat') or '').strip()
    email = (request.args.get('email') or '').strip()
    exclude = request.args.get('inspector_id')
    try:
        exclude_id = int(exclude) if exclude else None
    except (ValueError, TypeError):
        exclude_id = None
    if not nom or not prenom or not etat:
        return jsonify({'match': None})
    db = get_db()
    try:
        m = _find_formateur_match(db, nom, prenom, etat, email or None, exclude_id)
        return jsonify({'match': m})
    finally:
        db.close()

@app.route('/api/inspectors', methods=['POST'])
@auth_required
def add_inspector():
    if request.user['role'] not in ('National 1', 'Régional', 'Administrateur'):
        return jsonify({'error': 'Non autorisé'}), 403

    nom = request.form.get('nom', '')
    prenom = request.form.get('prenom', '')
    etat = request.form.get('etat', '')
    email = request.form.get('email', '')
    telephone = request.form.get('telephone', '')
    qualifications_json = request.form.get('qualifications', '[]')
    cv_file = request.files.get('cv')

    if not nom or not prenom or not etat:
        return jsonify({'error': 'Nom, prénom et État sont requis'}), 400

    db = get_db()
    try:
        # National 1 state restriction
        if request.user['role'] == 'National 1' and request.user.get('inspector_id'):
            user_ins = db.execute("SELECT etat FROM inspectors WHERE id = ?", (request.user['inspector_id'],)).fetchone()
            if user_ins and user_ins['etat'] != etat:
                return jsonify({'error': "Vous ne pouvez ajouter que des inspecteurs de votre État"}), 403

        ref = generate_reference(db)
        cv_path = None
        if cv_file and cv_file.filename:
            ext = os.path.splitext(cv_file.filename)[1]
            cv_path = f"cv-{secrets.token_hex(8)}{ext}"
            cv_file.save(os.path.join(UPLOAD_DIR, cv_path))

        cursor = db.execute("INSERT INTO inspectors (reference, nom, prenom, etat, email, telephone, cv_path) VALUES (?, ?, ?, ?, ?, ?, ?)",
            (ref, nom, prenom, etat, email or None, telephone or None, cv_path))
        inspector_id = cursor.lastrowid

        quals = json.loads(qualifications_json)
        for q in quals:
            if q.get('domaine'):
                db.execute("INSERT INTO qualifications (inspector_id, domaine, specialite, niveau, experience, titularisation) VALUES (?, ?, ?, ?, ?, ?)",
                    (inspector_id, q['domaine'], q.get('specialite', ''), q.get('niveau', ''), q.get('experience', ''), q.get('titularisation', '')))

        # Create user account with random password
        if email:
            chars = string.ascii_letters + string.digits
            auto_pw = ''.join(random.choices(chars, k=10))
            pw = bcrypt.hashpw(auto_pw.encode(), bcrypt.gensalt()).decode()
            try:
                db.execute("INSERT INTO users (username, password, role, must_change_password, inspector_id) VALUES (?, ?, 'National 2', 1, ?)",
                    (email, pw, inspector_id))
                log_activity(db, request.user['id'], 'CREATE_USER', f"Création auto utilisateur: {email} (MdP: {auto_pw})")
            except Exception:
                pass

        # Lien formateur (Option 2)
        try:
            fid, action = _handle_formateur_link(db, inspector_id, nom, prenom, etat, email or None, telephone or None)
            if action in ('linked', 'created'):
                log_activity(db, request.user['id'], 'LINK_FORMATEUR', f"Inspecteur #{inspector_id} {'créé en tant que' if action=='created' else 'lié au'} formateur #{fid}")
        except Exception as _e:
            print(f"[link formateur] {_e}")

        log_activity(db, request.user['id'], 'ADD_INSPECTOR', f"Ajout: {nom} {prenom} ({etat})")
        db.commit()
        return jsonify({'id': inspector_id, 'reference': ref, 'message': 'Inspecteur ajouté avec succès'})
    except Exception as e:
        db.rollback()
        return jsonify({'error': str(e)}), 500
    finally:
        db.close()

@app.route('/api/inspectors/<int:id>', methods=['PUT'])
@auth_required
def update_inspector(id):
    db = get_db()
    try:
        ins = db.execute("SELECT * FROM inspectors WHERE id = ?", (id,)).fetchone()
        if not ins:
            return jsonify({'error': 'Non trouvé'}), 404

        role = request.user['role']
        if role == 'National 2':
            if request.user.get('inspector_id') != id:
                return jsonify({'error': 'Vous ne pouvez modifier que vos propres informations'}), 403
            if request.form.get('email') and request.form['email'] != ins['email']:
                return jsonify({'error': "Vous ne pouvez pas modifier votre email"}), 403
        elif role == 'National 1':
            if request.user.get('inspector_id'):
                user_ins = db.execute("SELECT etat FROM inspectors WHERE id = ?", (request.user['inspector_id'],)).fetchone()
                if user_ins and user_ins['etat'] != ins['etat']:
                    return jsonify({'error': "Vous ne pouvez modifier que les inspecteurs de votre État"}), 403
        elif role not in ('Régional', 'Administrateur'):
            return jsonify({'error': 'Non autorisé'}), 403

        nom = request.form.get('nom', ins['nom'])
        prenom = request.form.get('prenom', ins['prenom'])
        etat = request.form.get('etat', ins['etat'])
        email = request.form.get('email', ins['email'])
        telephone = request.form.get('telephone', ins['telephone'])
        cv_file = request.files.get('cv')
        cv_path = ins['cv_path']
        if cv_file and cv_file.filename:
            ext = os.path.splitext(cv_file.filename)[1]
            cv_path = f"cv-{secrets.token_hex(8)}{ext}"
            cv_file.save(os.path.join(UPLOAD_DIR, cv_path))

        db.execute("UPDATE inspectors SET nom=?, prenom=?, etat=?, email=?, telephone=?, cv_path=?, updated_at=datetime('now') WHERE id=?",
            (nom, prenom, etat, email, telephone, cv_path, id))

        qualifications_json = request.form.get('qualifications')
        if qualifications_json:
            db.execute("DELETE FROM qualifications WHERE inspector_id = ?", (id,))
            for q in json.loads(qualifications_json):
                if q.get('domaine'):
                    db.execute("INSERT INTO qualifications (inspector_id, domaine, specialite, niveau, experience, titularisation) VALUES (?, ?, ?, ?, ?, ?)",
                        (id, q['domaine'], q.get('specialite', ''), q.get('niveau', ''), q.get('experience', ''), q.get('titularisation', '')))

        # Lien formateur (Option 2) — uniquement si admin/régional/national 1
        if request.user['role'] in ('National 1', 'Régional', 'Administrateur') and 'is_formateur' in request.form:
            try:
                fid, action = _handle_formateur_link(db, id, nom, prenom, etat, email or None, telephone or None)
                if action != 'none':
                    log_activity(db, request.user['id'], 'LINK_FORMATEUR', f"Inspecteur #{id}: {action} formateur #{fid or ''}")
            except Exception as _e:
                print(f"[link formateur] {_e}")

        log_activity(db, request.user['id'], 'UPDATE_INSPECTOR', f"Modification: {ins['reference']}")
        db.commit()
        return jsonify({'message': 'Inspecteur modifié avec succès'})
    except Exception as e:
        db.rollback()
        return jsonify({'error': str(e)}), 500
    finally:
        db.close()

@app.route('/api/inspectors/<int:id>/create-user', methods=['POST'])
@auth_required
def create_user_from_inspector(id):
    if request.user['role'] != 'Administrateur':
        return jsonify({'error': 'Non autorisé'}), 403
    db = get_db()
    try:
        ins = db.execute("SELECT id, nom, prenom, email, etat FROM inspectors WHERE id = ?", (id,)).fetchone()
        if not ins:
            return jsonify({'error': 'Inspecteur introuvable'}), 404
        if not ins['email']:
            return jsonify({'error': "L'email de l'inspecteur n'est pas renseigné. Veuillez d'abord renseigner l'email."}), 400
        existing = db.execute("SELECT id FROM users WHERE inspector_id = ? OR username = ?", (id, ins['email'])).fetchone()
        if existing:
            return jsonify({'error': 'Un compte utilisateur existe déjà pour cet inspecteur ou cet email'}), 400
        chars = string.ascii_letters + string.digits
        auto_pw = ''.join(random.choices(chars, k=10))
        pw = bcrypt.hashpw(auto_pw.encode(), bcrypt.gensalt()).decode()
        db.execute("INSERT INTO users (username, password, role, must_change_password, inspector_id, user_nom, user_prenom, user_etat) VALUES (?, ?, 'National 2', 1, ?, ?, ?, ?)",
                   (ins['email'], pw, id, ins['nom'], ins['prenom'], ins['etat']))
        log_activity(db, request.user['id'], 'CREATE_USER', f"Création compte utilisateur: {ins['email']} (MdP: {auto_pw})")
        db.commit()
        return jsonify({'message': 'Compte utilisateur créé avec succès', 'username': ins['email'], 'password': auto_pw})
    except Exception as e:
        db.rollback()
        return jsonify({'error': str(e)}), 500
    finally:
        db.close()

@app.route('/api/inspectors/<int:id>/deactivate', methods=['PUT'])
@auth_required
def deactivate_inspector(id):
    if request.user['role'] != 'Administrateur':
        return jsonify({'error': 'Non autorisé'}), 403
    db = get_db()
    db.execute("UPDATE inspectors SET is_active = 0, updated_at = datetime('now') WHERE id = ?", (id,))
    db.execute("UPDATE users SET is_active = 0 WHERE inspector_id = ?", (id,))
    log_activity(db, request.user['id'], 'DEACTIVATE', f"Désactivation inspecteur #{id}")
    db.commit()
    db.close()
    return jsonify({'message': 'Inspecteur désactivé'})

@app.route('/api/inspectors/<int:id>/activate', methods=['PUT'])
@auth_required
def activate_inspector(id):
    if request.user['role'] != 'Administrateur':
        return jsonify({'error': 'Non autorisé'}), 403
    db = get_db()
    db.execute("UPDATE inspectors SET is_active = 1, updated_at = datetime('now') WHERE id = ?", (id,))
    db.execute("UPDATE users SET is_active = 1 WHERE inspector_id = ?", (id,))
    log_activity(db, request.user['id'], 'ACTIVATE', f"Réactivation inspecteur #{id}")
    db.commit()
    db.close()
    return jsonify({'message': 'Inspecteur réactivé'})

@app.route('/api/inspectors/bulk-delete', methods=['POST'])
@auth_required
def bulk_delete_inspectors():
    if request.user['role'] != 'Administrateur':
        return jsonify({'error': 'Non autorisé'}), 403
    ids = request.json.get('ids', [])
    if not ids:
        return jsonify({'error': 'Aucun inspecteur sélectionné'}), 400
    db = get_db()
    placeholders = ','.join(['?'] * len(ids))
    db.execute(f"DELETE FROM qualifications WHERE inspector_id IN ({placeholders})", ids)
    db.execute(f"UPDATE users SET is_active = 0 WHERE inspector_id IN ({placeholders})", ids)
    db.execute(f"DELETE FROM inspectors WHERE id IN ({placeholders})", ids)
    log_activity(db, request.user['id'], 'BULK_DELETE_INSPECTORS', f"Suppression de {len(ids)} inspecteur(s)")
    db.commit()
    db.close()
    return jsonify({'message': f'{len(ids)} inspecteur(s) supprimé(s) avec succès'})

def _build_import_template(title, headers, sample_rows, instructions):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Modele'
    n = len(headers)
    last_col = get_column_letter(n)
    # Title
    ws.merge_cells(f'A1:{last_col}1')
    c = ws['A1']; c.value = title
    c.font = Font(name='Calibri', bold=True, size=14, color='1A365D')
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 30
    # Instructions
    ws.merge_cells(f'A2:{last_col}2')
    c = ws['A2']; c.value = instructions
    c.font = Font(italic=True, size=9, color='666666')
    c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws.row_dimensions[2].height = 24
    # Empty row
    ws.row_dimensions[3].height = 8
    # Headers (row 4)
    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='1A365D', end_color='1A365D', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin = Border(left=Side(style='thin', color='D0D0D0'), right=Side(style='thin', color='D0D0D0'),
                  top=Side(style='thin', color='D0D0D0'), bottom=Side(style='thin', color='D0D0D0'))
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=ci, value=h)
        cell.font = header_font; cell.fill = header_fill; cell.alignment = header_align; cell.border = thin
    ws.row_dimensions[4].height = 32
    # Sample rows (start row 5 — l'import lit à partir de la ligne 5)
    for ri, row in enumerate(sample_rows, 5):
        for ci, v in enumerate(row, 1):
            cell = ws.cell(row=ri, column=ci, value=v)
            cell.border = thin
    # Column widths
    for ci in range(1, n + 1):
        ws.column_dimensions[get_column_letter(ci)].width = 22
    out = io.BytesIO(); wb.save(out); out.seek(0)
    return out

@app.route('/api/inspectors/import-template', methods=['GET'])
@auth_required
def import_template_inspectors():
    if request.user['role'] != 'Administrateur':
        return jsonify({'error': 'Non autorise'}), 403
    headers = ['Référence (laisser vide)', 'Nom *', 'Prénom *', 'État *', 'Email', 'Téléphone',
               'Domaine', 'Spécialité', 'Niveau', 'Expérience', 'Titularisation (AAAA-MM)']
    samples = [
        ('', 'DUPONT', 'Jean', 'Bénin', 'jean.dupont@example.com', '+229 90000000',
         'OPS', 'Pilote ATR-72', 'Inspecteur Senior', '10 ans', '2018-03'),
        ('', 'KONE', 'Awa', 'Côte d\'Ivoire', 'awa.kone@example.com', '+225 0700000000',
         'AIR', 'Contrôle technique', 'Inspecteur Stagiaire', '2 ans', ''),
    ]
    instr = ('Lignes 1-4 = en-tête (ne pas modifier). Données à partir de la ligne 5. '
             'Champs marqués * obligatoires. Titularisation au format AAAA-MM (vide pour stagiaires).')
    out = _build_import_template('Modèle d\'import - Inspecteurs', headers, samples, instr)
    return send_file(out, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name='modele_import_inspecteurs.xlsx')

@app.route('/api/formateurs/import-template', methods=['GET'])
@auth_required
def import_template_formateurs():
    if request.user['role'] != 'Administrateur':
        return jsonify({'error': 'Non autorise'}), 403
    headers = ['Référence (laisser vide)', 'Nom *', 'Prénom *', 'État *', 'Email', 'Téléphone',
               'Aussi inspecteur (Oui/Non)', 'Type compétence', 'Domaine compétence',
               'Formations délivrées (séparées par ;)', 'Formations développées (séparées par ;)']
    samples = [
        ('', 'TRAORE', 'Issa', 'Burkina Faso', 'issa.traore@example.com', '+226 70000000',
         'Oui', 'Formateur National', 'OPS', 'Sécurité aérienne 2024;CRM avancé', 'Module Facteurs Humains'),
        ('', 'MENSAH', 'Adjoa', 'Togo', 'adjoa.mensah@example.com', '+228 90000000',
         'Non', 'Formateur Régional', 'AIR', '', ''),
    ]
    instr = ('Lignes 1-4 = en-tête (ne pas modifier). Données à partir de la ligne 5. '
             'Champs marqués * obligatoires. Plusieurs formations séparées par un point-virgule (;).')
    out = _build_import_template('Modèle d\'import - Formateurs', headers, samples, instr)
    return send_file(out, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name='modele_import_formateurs.xlsx')

@app.route('/api/inspectors/import', methods=['POST'])
@auth_required
def import_inspectors():
    if request.user['role'] != 'Administrateur':
        return jsonify({'error': 'Non autorisé'}), 403
    file = request.files.get('file')
    if not file:
        return jsonify({'error': 'Fichier requis'}), 400
    try:
        wb = openpyxl.load_workbook(file)
        ws = wb.active
        db = get_db()
        count = 0
        for row in ws.iter_rows(min_row=5, values_only=True):
            if not row or not row[1]:
                continue
            ref_val = row[0] or ''
            nom = str(row[1] or '').strip()
            prenom = str(row[2] or '').strip()
            etat = str(row[3] or '').strip()
            email = str(row[4] or '').strip()
            telephone = str(row[5] or '').strip()
            domaine = str(row[6] or '').strip()
            specialite = str(row[7] or '').strip()
            niveau = str(row[8] or '').strip()
            experience = str(row[9] or '').strip()
            titularisation = ''
            if len(row) > 10 and row[10]:
                tv = row[10]
                if hasattr(tv, 'strftime'):
                    titularisation = tv.strftime('%Y-%m')
                else:
                    titularisation = str(tv).strip()[:7]
            if not nom or not etat:
                continue
            existing = db.execute("SELECT id FROM inspectors WHERE nom = ? AND prenom = ? AND etat = ?", (nom, prenom, etat)).fetchone()
            if existing:
                ins_id = existing['id']
                db.execute("UPDATE inspectors SET email = ?, telephone = ?, updated_at = datetime('now') WHERE id = ?", (email, telephone, ins_id))
            else:
                ref = generate_reference(db)
                db.execute("INSERT INTO inspectors (reference, nom, prenom, etat, email, telephone) VALUES (?, ?, ?, ?, ?, ?)",
                           (ref, nom, prenom, etat, email, telephone))
                ins_id = db.lastrowid
                if email:
                    default_pw = '12345'
                    hashed = bcrypt.hashpw(default_pw.encode(), bcrypt.gensalt()).decode()
                    db.execute("INSERT OR IGNORE INTO users (username, password, role, inspector_id, must_change_password) VALUES (?, ?, 'National 2', ?, 1)",
                               (email, hashed, ins_id))
                count += 1
            if domaine:
                db.execute("INSERT INTO qualifications (inspector_id, domaine, specialite, niveau, experience, titularisation) VALUES (?, ?, ?, ?, ?, ?)",
                           (ins_id, domaine, specialite, niveau, experience, titularisation))
        log_activity(db, request.user['id'], 'IMPORT', f"Import Excel: {count} nouveaux inspecteurs")
        db.commit()
        db.close()
        return jsonify({'message': f'{count} inspecteur(s) importé(s) avec succès'})
    except Exception as e:
        return jsonify({'error': f"Erreur d'import: {str(e)}"}), 400

@app.route('/api/inspectors/<int:id>/email', methods=['POST'])
@auth_required
def send_email(id):
    if request.user['role'] not in ('Régional', 'Administrateur'):
        return jsonify({'error': 'Non autorisé'}), 403
    db = get_db()
    ins = db.execute("SELECT * FROM inspectors WHERE id = ?", (id,)).fetchone()
    if not ins or not ins['email']:
        db.close()
        return jsonify({'error': "Pas d'email"}), 400
    data = request.json
    subject = data.get('subject', '')
    body = data.get('body', '')
    log_activity(db, request.user['id'], 'SEND_EMAIL', f"Email à {ins['email']}: {subject}")
    db.commit()
    db.close()
    # Return mailto link for client-side email sending
    mailto = f"mailto:{ins['email']}?subject={urllib.parse.quote(subject)}&body={urllib.parse.quote(body)}"
    return jsonify({'message': f"Ouverture du client email pour {ins['email']}", 'mailto': mailto})

# ===== EXPORT HELPER =====
def get_export_data():
    """Get filtered inspector data for export"""
    db = get_db()
    where = ['i.is_active = 1']
    params = []
    etat = request.args.get('etat', '')
    domaine = request.args.get('domaine', '')
    status = request.args.get('status', 'active')

    if status == 'active':
        where = ['i.is_active = 1']
    elif status == 'inactive':
        where = ['i.is_active = 0']
    else:
        where = ['1=1']

    if etat: where.append('i.etat = ?'); params.append(etat)

    join = 'LEFT JOIN qualifications q ON q.inspector_id = i.id'
    if domaine: where.append('q.domaine = ?'); params.append(domaine)

    sql = f"SELECT i.reference, i.nom, i.prenom, i.etat, i.email, i.telephone, i.is_active, q.domaine, q.specialite, q.niveau, q.experience FROM inspectors i {join} WHERE {' AND '.join(where)} ORDER BY i.etat, i.nom"
    rows = db.execute(sql, params).fetchall()
    db.close()
    return rows

EXPORT_HEADERS = ['R\u00e9f\u00e9rence', 'Nom', 'Pr\u00e9nom', '\u00c9tat', 'Email', 'T\u00e9l\u00e9phone', 'Domaine', 'Sp\u00e9cialit\u00e9', 'Niveau', 'Exp\u00e9rience', 'Statut']

def row_to_list(r):
    return [r['reference'], r['nom'], r['prenom'], r['etat'], r['email'] or '', r['telephone'] or '', r['domaine'] or '', r['specialite'] or '', r['niveau'] or '', r['experience'] or '', 'Actif' if r['is_active'] else 'Inactif']

@app.route('/api/inspectors/export/csv', methods=['GET'])
@auth_required
def export_csv():
    rows = get_export_data()
    output = io.StringIO()
    output.write('\ufeff')
    writer = csv.writer(output, delimiter=';')
    writer.writerow(EXPORT_HEADERS)
    for r in rows:
        writer.writerow(row_to_list(r))
    output.seek(0)
    return send_file(io.BytesIO(output.getvalue().encode('utf-8-sig')),
        mimetype='text/csv', as_attachment=True, download_name='inspecteurs.csv')

@app.route('/api/inspectors/export/excel', methods=['GET'])
@auth_required
def export_excel():
    rows = get_export_data()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Inspecteurs'

    # Header style
    header_font = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='1A365D', end_color='1A365D', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin', color='D0D0D0'),
        right=Side(style='thin', color='D0D0D0'),
        top=Side(style='thin', color='D0D0D0'),
        bottom=Side(style='thin', color='D0D0D0')
    )

    # Title row
    ws.merge_cells('A1:K1')
    title_cell = ws['A1']
    title_cell.value = 'UEMOA - Base de donn\u00e9es des Inspecteurs et formateurs'
    title_cell.font = Font(name='Calibri', bold=True, size=14, color='1A365D')
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 35

    # Date row
    ws.merge_cells('A2:K2')
    date_cell = ws['A2']
    date_cell.value = 'Export\u00e9 le ' + datetime.now().strftime(_DT_FMT)
    date_cell.font = Font(name='Calibri', italic=True, size=10, color='666666')
    date_cell.alignment = Alignment(horizontal='center')
    ws.row_dimensions[2].height = 22

    # Empty row
    ws.row_dimensions[3].height = 8

    # Headers row (row 4)
    headers = EXPORT_HEADERS
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    ws.row_dimensions[4].height = 30

    # Data rows
    alt_fill = PatternFill(start_color='F7FAFC', end_color='F7FAFC', fill_type='solid')
    for row_idx, r in enumerate(rows, 5):
        data = row_to_list(r)
        for col_idx, value in enumerate(data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = Font(name='Calibri', size=10)
            cell.border = thin_border
            cell.alignment = Alignment(vertical='center', wrap_text=True)
            if (row_idx - 5) % 2 == 1:
                cell.fill = alt_fill
            # Color status cell
            if col_idx == 11:
                if value == 'Actif':
                    cell.font = Font(name='Calibri', size=10, color='166534', bold=True)
                else:
                    cell.font = Font(name='Calibri', size=10, color='991B1B', bold=True)

    # Auto-fit column widths
    col_widths = [16, 18, 18, 16, 28, 18, 10, 35, 22, 14, 10]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    # Freeze panes
    ws.freeze_panes = 'A5'

    # Auto-filter
    ws.auto_filter.ref = f'A4:K{4 + len(rows)}'

    # Summary sheet
    ws2 = wb.create_sheet('R\u00e9sum\u00e9')
    ws2.merge_cells('A1:C1')
    ws2['A1'].value = 'R\u00e9sum\u00e9 statistique'
    ws2['A1'].font = Font(name='Calibri', bold=True, size=14, color='1A365D')
    ws2['A1'].alignment = Alignment(horizontal='center')
    ws2.row_dimensions[1].height = 35

    ws2['A3'].value = 'Total inspecteurs export\u00e9s'
    ws2['A3'].font = Font(bold=True)
    ws2['B3'].value = len(rows)
    ws2['B3'].font = Font(bold=True, size=14, color='1A365D')

    ws2.column_dimensions['A'].width = 30
    ws2.column_dimensions['B'].width = 15
    ws2.column_dimensions['C'].width = 15

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True, download_name='inspecteurs.xlsx')

@app.route('/api/inspectors/export/pdf', methods=['GET'])
@auth_required
def export_pdf():
    rows = get_export_data()
    etat_filter = request.args.get('etat', '')
    domaine_filter = request.args.get('domaine', '')

    # Sanitize text for PDF
    def clean(text):
        if not text:
            return ''
        text = str(text)
        replacements = {
            '\u2019': "'", '\u2018': "'", '\u201c': '"', '\u201d': '"',
            '\u2013': '-', '\u2014': '-', '\u2026': '...', '\u00a0': ' ',
            '\u2022': '-', '\u200b': '', '\u200e': '', '\u200f': '',
            '\u0300': '', '\u0301': '', '\u0302': '', '\u0303': '', '\u0327': '',
            '\u2010': '-', '\u2011': '-', '\u2012': '-', '\u2015': '-',
            '\u2032': "'", '\u2033': '"', '\u02bc': "'", '\u02bb': "'",
            '\ufeff': '', '\u200c': '', '\u200d': '', '\u202a': '', '\u202c': '',
        }
        for old_c, new_c in replacements.items():
            text = text.replace(old_c, new_c)
        # Catch-all: replace any remaining non-latin1 characters
        try:
            text.encode('latin-1')
        except UnicodeEncodeError:
            text = text.encode('latin-1', 'replace').decode('latin-1')
        return text

    FONT_DIR = 'C:/Windows/Fonts' if os.path.exists('C:/Windows/Fonts') else '/usr/share/fonts/truetype/dejavu'

    class InspectorsPDF(FPDF):
        def __init__(self, *args, **kwargs):
            super().__init__(*args, **kwargs)
            if os.path.exists(os.path.join(FONT_DIR, 'arial.ttf')):
                self.add_font('ArialUni', '', os.path.join(FONT_DIR, 'arial.ttf'))
                self.add_font('ArialUni', 'B', os.path.join(FONT_DIR, 'arialbd.ttf'))
                self.add_font('ArialUni', 'I', os.path.join(FONT_DIR, 'ariali.ttf'))
            elif os.path.exists(os.path.join(FONT_DIR, 'DejaVuSans.ttf')):
                self.add_font('ArialUni', '', os.path.join(FONT_DIR, 'DejaVuSans.ttf'))
                self.add_font('ArialUni', 'B', os.path.join(FONT_DIR, 'DejaVuSans-Bold.ttf'))
                self.add_font('ArialUni', 'I', os.path.join(FONT_DIR, 'DejaVuSans-Oblique.ttf'))

        def header(self):
            self.set_font('ArialUni', 'B', 14)
            self.set_text_color(26, 54, 93)
            self.cell(0, 8, clean('UEMOA - Union \u00c9conomique et Mon\u00e9taire Ouest Africaine'), new_x="LMARGIN", new_y="NEXT", align='C')
            subtitle = "Base de donnees des Inspecteurs et formateurs"
            if etat_filter: subtitle += f' - Etat: {clean(etat_filter)}'
            if domaine_filter: subtitle += f' - Domaine: {clean(domaine_filter)}'
            self.set_font('ArialUni', '', 10)
            self.set_text_color(43, 87, 151)
            self.cell(0, 6, subtitle, new_x="LMARGIN", new_y="NEXT", align='C')
            self.set_font('ArialUni', 'I', 8)
            self.set_text_color(113, 128, 150)
            self.cell(0, 5, clean('Export\u00e9 le ' + datetime.now().strftime(_DT_FMT) + f' | Total: {len(rows)} enregistrement(s)'), new_x="LMARGIN", new_y="NEXT", align='C')
            self.set_draw_color(26, 54, 93)
            self.set_line_width(0.5)
            self.line(self.l_margin, self.get_y() + 2, self.w - self.r_margin, self.get_y() + 2)
            self.ln(5)

        def footer(self):
            self.set_y(-12)
            self.set_font('ArialUni', 'I', 7)
            self.set_text_color(113, 128, 150)
            self.cell(0, 10, clean(f"UEMOA - Direction de la Securite de l'Aviation Civile | Page {self.page_no()}/{{nb}}"), align='C')

    pdf = InspectorsPDF(orientation='L', format='A4')
    pdf.alias_nb_pages()
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.add_page()

    # Column config: name, width
    cols = [
        ('R\u00e9f.', 22),
        ('Nom', 28),
        ('Pr\u00e9nom', 28),
        ('\u00c9tat', 22),
        ('Email', 38),
        ('T\u00e9l.', 22),
        ('Dom.', 12),
        ('Sp\u00e9cialit\u00e9', 52),
        ('Niveau', 30),
        ('Exp.', 16),
        ('Statut', 14),
    ]

    # Table header
    pdf.set_font('ArialUni', 'B', 7)
    pdf.set_fill_color(26, 54, 93)
    pdf.set_text_color(255, 255, 255)
    for name, w in cols:
        pdf.cell(w, 7, name, border=1, fill=True, align='C')
    pdf.ln()

    # Table rows
    pdf.set_font('ArialUni', '', 6.5)
    fill = False
    for r in rows:
        if pdf.get_y() > pdf.h - 20:
            pdf.add_page()
            pdf.set_font('ArialUni', 'B', 7)
            pdf.set_fill_color(26, 54, 93)
            pdf.set_text_color(255, 255, 255)
            for name, w in cols:
                pdf.cell(w, 7, name, border=1, fill=True, align='C')
            pdf.ln()
            pdf.set_font('ArialUni', '', 6.5)
            fill = False

        if fill:
            pdf.set_fill_color(247, 250, 252)
        else:
            pdf.set_fill_color(255, 255, 255)

        pdf.set_text_color(26, 32, 44)
        data = [
            clean(r['reference']),
            clean(r['nom']),
            clean(r['prenom']),
            clean(r['etat']),
            clean(r['email']),
            clean(r['telephone']),
            clean(r['domaine']),
            clean((r['specialite'] or '')[:45]),
            clean((r['niveau'] or '')[:25]),
            clean(r['experience']),
            'Actif' if r['is_active'] else 'Inactif',
        ]

        for i, (name, w) in enumerate(cols):
            pdf.cell(w, 6, data[i], border='TB', fill=True)
        pdf.ln()
        fill = not fill

    output = io.BytesIO()
    pdf.output(output)
    output.seek(0)
    return send_file(output, mimetype='application/pdf', as_attachment=True, download_name='inspecteurs.pdf')


# ===== ANALYTICS ROUTES =====
def _compute_analytics(db, f_etat, f_domaine, f_niveau, f_competence):

    # --- Inspector filter conditions (qual-joined queries) ---
    q_conds  = ["i.is_active = 1"]
    q_params = []
    if f_etat:    q_conds.append("i.etat = ?");         q_params.append(f_etat)
    if f_domaine: q_conds.append("q.domaine = ?");       q_params.append(f_domaine)
    if f_niveau:  q_conds.append("TRIM(q.niveau) = ?");  q_params.append(f_niveau)
    q_where = " AND ".join(q_conds)

    # by_state: no qual join unless domaine/niveau filter active
    if f_domaine or f_niveau:
        by_state = [dict(r) for r in db.execute(
            "SELECT i.etat, COUNT(DISTINCT i.id) as count FROM qualifications q"
            " JOIN inspectors i ON q.inspector_id=i.id"
            " WHERE " + q_where + " GROUP BY i.etat ORDER BY count DESC",
            q_params).fetchall()]
    else:
        s_conds  = ["is_active = 1"]
        s_params = []
        if f_etat: s_conds.append("etat = ?"); s_params.append(f_etat)
        by_state = [dict(r) for r in db.execute(
            "SELECT etat, COUNT(*) as count FROM inspectors WHERE " +
            " AND ".join(s_conds) + " GROUP BY etat ORDER BY count DESC",
            s_params).fetchall()]

    by_domain = [dict(r) for r in db.execute(
        "SELECT q.domaine, COUNT(DISTINCT q.inspector_id) as count"
        " FROM qualifications q JOIN inspectors i ON q.inspector_id = i.id"
        " WHERE " + q_where + " GROUP BY q.domaine ORDER BY count DESC",
        q_params).fetchall()]

    by_level = [dict(r) for r in db.execute(
        "SELECT TRIM(q.niveau) as niveau, COUNT(*) as count"
        " FROM qualifications q JOIN inspectors i ON q.inspector_id = i.id"
        " WHERE " + q_where + " GROUP BY TRIM(q.niveau) ORDER BY count DESC",
        q_params).fetchall()]

    by_exp = [dict(r) for r in db.execute(
        "SELECT q.experience, COUNT(*) as count"
        " FROM qualifications q JOIN inspectors i ON q.inspector_id = i.id"
        " WHERE " + q_where + " AND q.experience != '' AND q.experience IS NOT NULL"
        " GROUP BY q.experience ORDER BY count DESC",
        q_params).fetchall()]

    domain_state = [dict(r) for r in db.execute(
        "SELECT i.etat, q.domaine, COUNT(DISTINCT q.inspector_id) as count"
        " FROM qualifications q JOIN inspectors i ON q.inspector_id = i.id"
        " WHERE " + q_where + " GROUP BY i.etat, q.domaine ORDER BY i.etat, q.domaine",
        q_params).fetchall()]

    by_speciality = [dict(r) for r in db.execute(
        "SELECT SUBSTR(q.specialite, 1, 40) as specialite, COUNT(*) as count"
        " FROM qualifications q JOIN inspectors i ON q.inspector_id = i.id"
        " WHERE " + q_where + " GROUP BY SUBSTR(q.specialite, 1, 40) ORDER BY count DESC LIMIT 15",
        q_params).fetchall()]

    # Niveaux par Domaines (cross-tab)
    level_domain = [dict(r) for r in db.execute(
        "SELECT q.domaine, TRIM(q.niveau) as niveau, COUNT(DISTINCT q.inspector_id) as count"
        " FROM qualifications q JOIN inspectors i ON q.inspector_id = i.id"
        " WHERE " + q_where + " GROUP BY q.domaine, TRIM(q.niveau) ORDER BY q.domaine, TRIM(q.niveau)",
        q_params).fetchall()]

    # --- Formateurs analytics ---
    # Conditions pour requêtes simples (sans JOIN, pas d'alias)
    fs_conds  = ["is_active = 1"]
    fs_params = []
    if f_etat: fs_conds.append("etat = ?"); fs_params.append(f_etat)
    fs_where = " AND ".join(fs_conds)
    # Conditions pour requêtes avec JOIN formateur_competences (alias f)
    fj_conds  = ["f.is_active = 1"]
    fj_params = []
    if f_etat: fj_conds.append("f.etat = ?"); fj_params.append(f_etat)
    if f_competence: fj_conds.append("fc.type_competence = ?"); fj_params.append(f_competence)
    fj_where = " AND ".join(fj_conds)

    if f_competence:
        frm_total = db.execute(
            "SELECT COUNT(DISTINCT f.id) as count"
            " FROM formateur_competences fc JOIN formateurs f ON fc.formateur_id = f.id"
            " WHERE " + fj_where, fj_params).fetchone()['count']
        frm_by_state = [dict(r) for r in db.execute(
            "SELECT f.etat, COUNT(DISTINCT f.id) as count"
            " FROM formateur_competences fc JOIN formateurs f ON fc.formateur_id = f.id"
            " WHERE " + fj_where + " GROUP BY f.etat ORDER BY count DESC",
            fj_params).fetchall()]
        frm_inspecteurs = db.execute(
            "SELECT COUNT(DISTINCT f.id) as count"
            " FROM formateur_competences fc JOIN formateurs f ON fc.formateur_id = f.id"
            " WHERE " + fj_where + " AND f.is_inspecteur = 1",
            fj_params).fetchone()['count']
    else:
        frm_total       = db.execute("SELECT COUNT(*) as count FROM formateurs WHERE " + fs_where, fs_params).fetchone()['count']
        frm_by_state    = [dict(r) for r in db.execute("SELECT etat, COUNT(*) as count FROM formateurs WHERE " + fs_where + " GROUP BY etat ORDER BY count DESC", fs_params).fetchall()]
        frm_inspecteurs = db.execute("SELECT COUNT(*) as count FROM formateurs WHERE " + fs_where + " AND is_inspecteur = 1", fs_params).fetchone()['count']

    fc_cond   = fj_where
    fc_params = fj_params

    frm_by_competence = [dict(r) for r in db.execute(
        "SELECT fc.type_competence, COUNT(DISTINCT fc.formateur_id) as count"
        " FROM formateur_competences fc JOIN formateurs f ON fc.formateur_id = f.id"
        " WHERE " + fc_cond + " GROUP BY fc.type_competence ORDER BY count DESC",
        fc_params).fetchall()]

    frm_by_domaine = [dict(r) for r in db.execute(
        "SELECT fc.domaine, COUNT(DISTINCT fc.formateur_id) as count"
        " FROM formateur_competences fc JOIN formateurs f ON fc.formateur_id = f.id"
        " WHERE " + fc_cond + " AND fc.domaine IS NOT NULL AND fc.domaine != ''"
        " GROUP BY fc.domaine ORDER BY count DESC",
        fc_params).fetchall()]

    frm_comp_state = [dict(r) for r in db.execute(
        "SELECT f.etat, fc.type_competence, COUNT(DISTINCT fc.formateur_id) as count"
        " FROM formateur_competences fc JOIN formateurs f ON fc.formateur_id = f.id"
        " WHERE " + fc_cond + " GROUP BY f.etat, fc.type_competence ORDER BY f.etat",
        fc_params).fetchall()]

    return {
        'byState': by_state, 'byDomain': by_domain, 'byLevel': by_level,
        'byExperience': by_exp, 'domainState': domain_state, 'bySpeciality': by_speciality,
        'levelDomain': level_domain,
        'formateurs': {
            'total': frm_total,
            'byState': frm_by_state,
            'byCompetence': frm_by_competence,
            'byDomaine': frm_by_domaine,
            'inspecteurs': frm_inspecteurs,
            'competenceState': frm_comp_state
        }
    }

@app.route('/api/analytics', methods=['GET'])
@auth_required
def analytics_data():
    if request.user.get('role') == 'National 2':
        return jsonify({'error': 'Non autorise'}), 403
    db = get_db()
    try:
        payload = _compute_analytics(db,
            request.args.get('etat', '').strip(),
            request.args.get('domaine', '').strip(),
            request.args.get('niveau', '').strip(),
            request.args.get('competence', '').strip())
        return jsonify(payload)
    finally:
        db.close()

@app.route('/api/analytics/export/<fmt>', methods=['GET'])
@auth_required
def export_analytics(fmt):
    if request.user.get('role') == 'National 2':
        return jsonify({'error': 'Non autorise'}), 403
    if fmt not in ('excel', 'pdf'):
        return jsonify({'error': 'Format invalide'}), 400
    db = get_db()
    try:
        d = _compute_analytics(db,
            request.args.get('etat', '').strip(),
            request.args.get('domaine', '').strip(),
            request.args.get('niveau', '').strip(),
            request.args.get('competence', '').strip())
    finally:
        db.close()

    sections = [
        ('Inspecteurs par État',      ['État', 'Nombre'],     [(r['etat'], r['count']) for r in d['byState']]),
        ('Inspecteurs par Domaine',   ['Domaine', 'Nombre'],  [(r['domaine'], r['count']) for r in d['byDomain']]),
        ('Répartition par Niveau',    ['Niveau', 'Nombre'],   [(r['niveau'], r['count']) for r in d['byLevel']]),
        ('Répartition par Expérience', ['Expérience', 'Nombre'], [(r['experience'], r['count']) for r in d['byExperience']]),
        ('Domaines par État',         ['État', 'Domaine', 'Nombre'], [(r['etat'], r['domaine'], r['count']) for r in d['domainState']]),
        ('Niveaux par Domaines',      ['Domaine', 'Niveau', 'Nombre'], [(r['domaine'], r['niveau'], r['count']) for r in d['levelDomain']]),
        ('Top 15 Spécialités',        ['Spécialité', 'Nombre'], [(r['specialite'], r['count']) for r in d['bySpeciality']]),
        ('Formateurs par État',       ['État', 'Nombre'],     [(r['etat'], r['count']) for r in d['formateurs']['byState']]),
        ('Formateurs par Compétence', ['Compétence', 'Nombre'], [(r['type_competence'], r['count']) for r in d['formateurs']['byCompetence']]),
        ('Formateurs par Domaine',    ['Domaine', 'Nombre'], [(r['domaine'], r['count']) for r in d['formateurs']['byDomaine']]),
        ('Compétences par État',      ['État', 'Compétence', 'Nombre'], [(r['etat'], r['type_competence'], r['count']) for r in d['formateurs']['competenceState']]),
    ]
    today = datetime.now().strftime('%Y%m%d_%H%M')

    if fmt == 'excel':
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        header_font = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
        header_fill = PatternFill(start_color='1A365D', end_color='1A365D', fill_type='solid')
        header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        thin = Border(left=Side(style='thin', color='D0D0D0'), right=Side(style='thin', color='D0D0D0'),
                      top=Side(style='thin', color='D0D0D0'), bottom=Side(style='thin', color='D0D0D0'))
        for title, headers, rows in sections:
            ws = wb.create_sheet(title=title[:31])
            ws.cell(row=1, column=1, value=title).font = Font(bold=True, size=13, color='1A365D')
            ws.cell(row=2, column=1, value='Exporté le ' + datetime.now().strftime(_DT_FMT)).font = Font(italic=True, size=9, color='666666')
            for ci, h in enumerate(headers, 1):
                c = ws.cell(row=4, column=ci, value=h)
                c.font = header_font; c.fill = header_fill; c.alignment = header_align; c.border = thin
            for ri, row in enumerate(rows, 5):
                for ci, val in enumerate(row, 1):
                    c = ws.cell(row=ri, column=ci, value=val)
                    c.border = thin
            for ci in range(1, len(headers) + 1):
                ws.column_dimensions[get_column_letter(ci)].width = 28 if ci == 1 else 16
        out = io.BytesIO(); wb.save(out); out.seek(0)
        return send_file(out, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                         as_attachment=True, download_name=f'statistiques_{today}.xlsx')

    # PDF
    pdf = FPDF(orientation='P', unit='mm', format='A4')
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.add_page()
    pdf.set_font('Helvetica', 'B', 16)
    pdf.cell(0, 10, 'UEMOA - Statistiques Inspecteurs et Formateurs', ln=1, align='C')
    pdf.set_font('Helvetica', 'I', 9)
    pdf.cell(0, 6, 'Exporté le ' + datetime.now().strftime(_DT_FMT), ln=1, align='C')
    pdf.ln(4)
    for title, headers, rows in sections:
        # Estimer hauteur nécessaire pour ne pas couper le tableau
        needed = 12 + 8 + 7 * (len(rows) + 1) + 6
        if pdf.get_y() + needed > pdf.h - 15:
            pdf.add_page()
        pdf.set_font('Helvetica', 'B', 12)
        pdf.set_text_color(26, 54, 93)
        pdf.cell(0, 8, title, ln=1)
        pdf.set_text_color(0, 0, 0)
        pdf.set_font('Helvetica', 'B', 9)
        pdf.set_fill_color(26, 54, 93); pdf.set_text_color(255, 255, 255)
        col_count = len(headers)
        avail = pdf.w - 30
        widths = [avail * 0.55] + [avail * 0.45 / (col_count - 1)] * (col_count - 1) if col_count > 1 else [avail]
        if col_count == 3:
            widths = [avail * 0.40, avail * 0.40, avail * 0.20]
        for h, w in zip(headers, widths):
            pdf.cell(w, 7, str(h), border=1, align='C', fill=True)
        pdf.ln()
        pdf.set_text_color(0, 0, 0); pdf.set_font('Helvetica', '', 9); pdf.set_fill_color(247, 250, 252)
        alt = False
        for row in rows:
            if pdf.get_y() + 7 > pdf.h - 15:
                pdf.add_page()
                pdf.set_font('Helvetica', 'B', 9)
                pdf.set_fill_color(26, 54, 93); pdf.set_text_color(255, 255, 255)
                for h, w in zip(headers, widths):
                    pdf.cell(w, 7, str(h), border=1, align='C', fill=True)
                pdf.ln()
                pdf.set_text_color(0, 0, 0); pdf.set_font('Helvetica', '', 9); pdf.set_fill_color(247, 250, 252)
            for v, w in zip(row, widths):
                txt = str(v) if v is not None else ''
                if len(txt) > 60: txt = txt[:57] + '...'
                try:
                    pdf.cell(w, 7, txt, border=1, fill=alt)
                except Exception:
                    pdf.cell(w, 7, txt.encode('latin-1', 'replace').decode('latin-1'), border=1, fill=alt)
            pdf.ln(); alt = not alt
        pdf.ln(4)
    raw = pdf.output(dest='S')
    if isinstance(raw, str):
        raw = raw.encode('latin-1')
    elif isinstance(raw, bytearray):
        raw = bytes(raw)
    out = io.BytesIO(raw); out.seek(0)
    return send_file(out, mimetype='application/pdf', as_attachment=True, download_name=f'statistiques_{today}.pdf')

# ===== ADMIN ROUTES =====
@app.route('/api/admin/users', methods=['GET'])
@auth_required
def admin_users():
    if request.user['role'] != 'Administrateur':
        return jsonify({'error': 'Non autorisé'}), 403
    db = get_db()
    users = [dict(r) for r in db.execute("""
        SELECT u.id, u.username, u.role, u.is_active, u.must_change_password, u.inspector_id, u.created_at,
               COALESCE(i.nom, u.user_nom) as nom, COALESCE(i.prenom, u.user_prenom) as prenom, COALESCE(i.etat, u.user_etat) as etat, i.email
        FROM users u LEFT JOIN inspectors i ON u.inspector_id = i.id ORDER BY u.role, u.username
    """).fetchall()]

    for u in users:
        last = db.execute("SELECT created_at FROM activity_log WHERE user_id = ? AND action = 'LOGIN' ORDER BY created_at DESC LIMIT 1", (u['id'],)).fetchone()
        u['lastLogin'] = last['created_at'] if last else None
        u['recentActions'] = [dict(a) for a in db.execute("SELECT action, details, created_at FROM activity_log WHERE user_id = ? ORDER BY created_at DESC LIMIT 5", (u['id'],)).fetchall()]

    db.close()
    return jsonify(users)

@app.route('/api/admin/users', methods=['POST'])
@auth_required
def create_user():
    if request.user['role'] != 'Administrateur':
        return jsonify({'error': 'Non autorisé'}), 403
    data = request.json
    username = data.get('username', '').strip()
    nom = data.get('nom', '').strip()
    etat = data.get('etat', '').strip()
    role = data.get('role', 'National 2')
    if not username:
        return jsonify({'error': "Nom d'utilisateur requis"}), 400
    if role not in ('National 2', 'National 1', 'Régional', 'Administrateur'):
        return jsonify({'error': 'Rôle invalide'}), 400
    # Auto-generate password
    chars = string.ascii_letters + string.digits
    password = data.get('password', '').strip()
    if not password:
        password = ''.join(random.choices(chars, k=10))
    if len(password) < 6:
        return jsonify({'error': 'Le mot de passe doit contenir au moins 6 caractères'}), 400
    db = get_db()
    try:
        existing = db.execute("SELECT id FROM users WHERE username = ?", (username,)).fetchone()
        if existing:
            return jsonify({'error': "Ce nom d'utilisateur existe déjà"}), 400
        hashed = bcrypt.hashpw(password.encode(), bcrypt.gensalt()).decode()
        db.execute("INSERT INTO users (username, password, role, must_change_password, user_nom, user_etat) VALUES (?, ?, ?, 1, ?, ?)",
                   (username, hashed, role, nom or None, etat or None))
        log_activity(db, request.user['id'], 'CREATE_USER', f"Création utilisateur: {username} ({role}) (MdP: {password})")
        db.commit()
        return jsonify({'message': f"Utilisateur '{username}' créé avec succès. Mot de passe: {password}", 'password': password})
    except Exception as e:
        db.rollback()
        return jsonify({'error': str(e)}), 500
    finally:
        db.close()

@app.route('/api/admin/users/<int:id>/role', methods=['PUT'])
@auth_required
def update_role(id):
    if request.user['role'] != 'Administrateur':
        return jsonify({'error': 'Non autorisé'}), 403
    role = request.json.get('role')
    if role not in ('National 2', 'National 1', 'Régional', 'Administrateur'):
        return jsonify({'error': 'Rôle invalide'}), 400
    db = get_db()
    db.execute("UPDATE users SET role = ?, updated_at = datetime('now') WHERE id = ?", (role, id))
    log_activity(db, request.user['id'], 'CHANGE_ROLE', f"Rôle #{id}: {role}")
    db.commit()
    db.close()
    return jsonify({'message': 'Rôle mis à jour'})

@app.route('/api/admin/users/<int:id>/username', methods=['PUT'])
@auth_required
def update_username(id):
    if request.user['role'] != 'Administrateur':
        return jsonify({'error': 'Non autorisé'}), 403
    username = request.json.get('username')
    if not username:
        return jsonify({'error': "Nom d'utilisateur requis"}), 400
    db = get_db()
    try:
        db.execute("UPDATE users SET username = ?, updated_at = datetime('now') WHERE id = ?", (username, id))
        db.commit()
    except Exception as e:
        if 'UNIQUE' in str(e).upper() or 'unique' in str(e).lower() or 'duplicate' in str(e).lower():
            db.close()
            return jsonify({'error': "Ce nom d'utilisateur existe déjà"}), 400
        db.close()
        return jsonify({'error': str(e)}), 500
    log_activity(db, request.user['id'], 'CHANGE_USERNAME', f"Username #{id}: {username}")
    db.commit()
    db.close()
    return jsonify({'message': "Nom d'utilisateur mis à jour"})

@app.route('/api/admin/users/<int:id>/info', methods=['PUT'])
@auth_required
def update_user_info(id):
    if request.user['role'] != 'Administrateur':
        return jsonify({'error': 'Non autorisé'}), 403
    data = request.json
    nom    = (data.get('nom') or '').strip()
    prenom = (data.get('prenom') or '').strip()
    etat   = (data.get('etat') or '').strip()
    db = get_db()
    db.execute("UPDATE users SET user_nom = ?, user_etat = ?, user_prenom = ?, updated_at = datetime('now') WHERE id = ?",
               (nom or None, etat or None, prenom or None, id))
    log_activity(db, request.user['id'], 'UPDATE_USER_INFO', f"Info #{id}: {nom} {prenom} / {etat}")
    db.commit()
    db.close()
    return jsonify({'message': 'Informations mises à jour'})

@app.route('/api/admin/users/<int:id>/reset-password', methods=['PUT'])
@auth_required
def reset_password(id):
    if request.user['role'] != 'Administrateur':
        return jsonify({'error': 'Non autorisé'}), 403
    db = get_db()
    user = db.execute("SELECT * FROM users WHERE id = ?", (id,)).fetchone()
    if not user:
        db.close()
        return jsonify({'error': 'Non trouvé'}), 404
    inspector = db.execute("SELECT etat, email FROM inspectors WHERE id = ?", (user['inspector_id'],)).fetchone() if user['inspector_id'] else None
    # Generate random password
    chars = string.ascii_letters + string.digits
    new_pw = ''.join(random.choices(chars, k=10))
    hashed = bcrypt.hashpw(new_pw.encode(), bcrypt.gensalt()).decode()
    # Store the plain text password temporarily for admin reveal
    db.execute("UPDATE users SET password = ?, must_change_password = 1, updated_at = datetime('now') WHERE id = ?", (hashed, id))
    # Store the plain password in a temporary field for reveal feature
    db.execute("UPDATE users SET inspector_id = inspector_id WHERE id = ?", (id,))  # no-op to trigger updated_at
    log_activity(db, request.user['id'], 'RESET_PASSWORD', f"Reset pour: {user['username']} (MdP: {new_pw})")
    db.commit()

    # Build mailto link for email notification
    recipient_email = inspector['email'] if inspector and inspector['email'] else (user['username'] if '@' in user['username'] else None)

    db.close()
    msg = f'Mot de passe réinitialisé. Nouveau: {new_pw}'
    return jsonify({
        'message': msg,
        'newPassword': new_pw,
        'recipientEmail': recipient_email,
        'username': user['username']
    })

@app.route('/api/admin/users/<int:id>/toggle-active', methods=['PUT'])
@auth_required
def toggle_active(id):
    if request.user['role'] != 'Administrateur':
        return jsonify({'error': 'Non autorisé'}), 403
    db = get_db()
    user = db.execute("SELECT * FROM users WHERE id = ?", (id,)).fetchone()
    if not user:
        db.close()
        return jsonify({'error': 'Non trouvé'}), 404
    new_status = 0 if user['is_active'] else 1
    db.execute("UPDATE users SET is_active = ?, updated_at = datetime('now') WHERE id = ?", (new_status, id))
    log_activity(db, request.user['id'], 'TOGGLE_USER', f"{'Activé' if new_status else 'Désactivé'}: {user['username']}")
    db.commit()
    db.close()
    return jsonify({'message': 'Utilisateur activé' if new_status else 'Utilisateur désactivé'})

@app.route('/api/admin/users/<int:id>/reveal-password', methods=['POST'])
@auth_required
def reveal_password(id):
    """Reveal user's default password after admin confirms their own password."""
    if request.user['role'] != 'Administrateur':
        return jsonify({'error': 'Non autorisé'}), 403
    admin_password = request.json.get('adminPassword', '')
    if not admin_password:
        return jsonify({'error': "Mot de passe administrateur requis"}), 400
    db = get_db()
    admin = db.execute("SELECT password FROM users WHERE id = ?", (request.user['id'],)).fetchone()
    if not admin or not bcrypt.checkpw(admin_password.encode(), admin['password'].encode()):
        db.close()
        return jsonify({'error': 'Mot de passe administrateur incorrect'}), 401
    user = db.execute("SELECT * FROM users WHERE id = ?", (id,)).fetchone()
    if not user:
        db.close()
        return jsonify({'error': 'Utilisateur non trouvé'}), 404
    # Try to find the last reset password from activity log
    last_reset = db.execute(
        "SELECT details FROM activity_log WHERE action = 'RESET_PASSWORD' AND details LIKE ? ORDER BY created_at DESC LIMIT 1",
        (f"Reset pour: {user['username']}%",)
    ).fetchone()
    found_pw = None
    if last_reset and '(MdP: ' in last_reset['details']:
        found_pw = last_reset['details'].split('(MdP: ')[-1].rstrip(')')
    if found_pw:
        # Verify this password is still the current one
        is_current = bcrypt.checkpw(found_pw.encode(), user['password'].encode())
        if is_current:
            log_activity(db, request.user['id'], 'REVEAL_PASSWORD', f"Consultation MdP de: {user['username']}")
            db.commit()
            db.close()
            return jsonify({'password': found_pw, 'isDefault': True})
    log_activity(db, request.user['id'], 'REVEAL_PASSWORD', f"Consultation MdP de: {user['username']}")
    db.commit()
    db.close()
    return jsonify({'password': None, 'isDefault': False, 'message': "L'utilisateur a modifié son mot de passe."})

@app.route('/api/admin/logs', methods=['GET'])
@auth_required
def get_logs():
    if request.user['role'] != 'Administrateur':
        return jsonify({'error': 'Non autorisé'}), 403
    db = get_db()
    user_id = request.args.get('userId')
    limit = int(request.args.get('limit', 100))
    if user_id:
        logs = [dict(r) for r in db.execute("SELECT al.*, u.username FROM activity_log al LEFT JOIN users u ON al.user_id = u.id WHERE al.user_id = ? ORDER BY al.created_at DESC LIMIT ?", (user_id, limit)).fetchall()]
    else:
        logs = [dict(r) for r in db.execute("SELECT al.*, u.username FROM activity_log al LEFT JOIN users u ON al.user_id = u.id ORDER BY al.created_at DESC LIMIT ?", (limit,)).fetchall()]
    db.close()
    return jsonify(logs)

@app.route('/api/admin/notifications', methods=['GET'])
@auth_required
def get_notifications():
    if request.user['role'] != 'Administrateur':
        return jsonify({'error': 'Non autorisé'}), 403
    db = get_db()
    alerts = [dict(r) for r in db.execute("SELECT * FROM activity_log WHERE action IN ('RESET_ALERT', 'RESET_NOTIFICATION') ORDER BY created_at DESC LIMIT 20").fetchall()]
    db.close()
    return jsonify({'alerts': alerts})

# ===== SETTINGS ROUTES =====
@app.route('/api/settings', methods=['GET'])
@auth_required
def get_settings():
    if request.user.get('role') not in ('Administrateur', 'Régional'):
        return jsonify({'error': 'Non autorisé'}), 403
    db = get_db()
    category = request.args.get('category', '')
    if category:
        rows = [dict(r) for r in db.execute("SELECT * FROM settings WHERE category = ? ORDER BY sort_order, value", (category,)).fetchall()]
    else:
        rows = [dict(r) for r in db.execute("SELECT * FROM settings ORDER BY category, sort_order, value").fetchall()]
    db.close()
    return jsonify(rows)

@app.route('/api/settings/active', methods=['GET'])
@auth_required
def get_active_settings():
    db = get_db()
    category = request.args.get('category', '')
    if category:
        rows = [dict(r) for r in db.execute("SELECT * FROM settings WHERE category = ? AND is_active = 1 ORDER BY sort_order, value", (category,)).fetchall()]
    else:
        rows = [dict(r) for r in db.execute("SELECT * FROM settings WHERE is_active = 1 ORDER BY category, sort_order, value").fetchall()]
    db.close()
    return jsonify(rows)

@app.route('/api/settings', methods=['POST'])
@auth_required
def add_setting():
    if request.user.get('role') != 'Administrateur':
        return jsonify({'error': 'Non autorisé'}), 403
    data = request.json
    category = data.get('category', '').strip()
    value = data.get('value', '').strip()
    label = data.get('label', '').strip() or None
    if not category or not value:
        return jsonify({'error': 'Catégorie et valeur requises'}), 400
    db = get_db()
    existing = db.execute("SELECT id FROM settings WHERE category = ? AND value = ?", (category, value)).fetchone()
    if existing:
        db.close()
        return jsonify({'error': 'Cette valeur existe déjà dans cette catégorie'}), 400
    db.execute("INSERT INTO settings (category, value, label) VALUES (?, ?, ?)", (category, value, label))
    log_activity(db, request.user['id'], 'ADD_SETTING', f"Paramètre ajouté: {category}/{value}")
    db.commit()
    db.close()
    return jsonify({'message': 'Paramètre ajouté'})

@app.route('/api/settings/<int:id>', methods=['PUT'])
@auth_required
def update_setting(id):
    if request.user.get('role') != 'Administrateur':
        return jsonify({'error': 'Non autorisé'}), 403
    data = request.json
    db = get_db()
    old = db.execute("SELECT * FROM settings WHERE id = ?", (id,)).fetchone()
    if not old:
        db.close()
        return jsonify({'error': 'Non trouvé'}), 404
    new_value = data.get('value', old['value']).strip()
    new_label = data.get('label', old['label'])
    if new_label:
        new_label = new_label.strip()
    old_value = old['value']
    cat = old['category']
    db.execute("UPDATE settings SET value = ?, label = ?, updated_at = datetime('now') WHERE id = ?",
               (new_value, new_label, id))
    if old_value != new_value:
        if cat == 'etat':
            db.execute("UPDATE inspectors SET etat = ? WHERE etat = ?", (new_value, old_value))
            db.execute("UPDATE formateurs SET etat = ? WHERE etat = ?", (new_value, old_value))
        elif cat == 'domaine':
            db.execute("UPDATE qualifications SET domaine = ? WHERE domaine = ?", (new_value, old_value))
            db.execute("UPDATE formateur_competences SET domaine = ? WHERE domaine = ?", (new_value, old_value))
        elif cat == 'niveau':
            db.execute("UPDATE qualifications SET niveau = ? WHERE niveau = ?", (new_value, old_value))
        elif cat == 'formateur':
            db.execute("UPDATE formateur_competences SET type_competence = ? WHERE type_competence = ?", (new_value, old_value))
    log_activity(db, request.user['id'], 'UPDATE_SETTING', f"Paramètre modifié: {cat}/{old_value} -> {new_value}")
    db.commit()
    db.close()
    return jsonify({'message': 'Paramètre modifié'})

@app.route('/api/settings/<int:id>/toggle', methods=['PUT'])
@auth_required
def toggle_setting(id):
    if request.user.get('role') != 'Administrateur':
        return jsonify({'error': 'Non autorisé'}), 403
    db = get_db()
    setting = db.execute("SELECT * FROM settings WHERE id = ?", (id,)).fetchone()
    if not setting:
        db.close()
        return jsonify({'error': 'Non trouvé'}), 404
    new_status = 0 if setting['is_active'] else 1
    db.execute("UPDATE settings SET is_active = ?, updated_at = datetime('now') WHERE id = ?", (new_status, id))
    log_activity(db, request.user['id'], 'TOGGLE_SETTING', f"{'Activé' if new_status else 'Désactivé'}: {setting['category']}/{setting['value']}")
    db.commit()
    db.close()
    return jsonify({'message': 'Activé' if new_status else 'Désactivé'})

@app.route('/api/settings/bulk-delete', methods=['POST'])
@auth_required
def bulk_delete_settings():
    if request.user['role'] != 'Administrateur':
        return jsonify({'error': 'Non autorisé'}), 403
    ids = request.json.get('ids', [])
    if not ids:
        return jsonify({'error': 'Aucun paramètre sélectionné'}), 400
    db = get_db()
    try:
        for sid in ids:
            setting = db.execute("SELECT * FROM settings WHERE id = ?", (sid,)).fetchone()
            if setting:
                db.execute("DELETE FROM settings WHERE id = ?", (sid,))
                log_activity(db, request.user['id'], 'DELETE_SETTING', f"Paramètre supprimé: {setting['category']}/{setting['value']}")
        db.commit()
    finally:
        db.close()
    return jsonify({'message': f'{len(ids)} paramètre(s) supprimé(s)'})


# ===== FORMATEURS ROUTES =====
def generate_formateur_reference(db):
    row = db.execute("SELECT reference FROM formateurs ORDER BY id DESC LIMIT 1").fetchone()
    num = 1
    if row:
        try: num = int(row['reference'].split('-')[-1]) + 1
        except: pass
    return f"UEMOA-FRM-{num:04d}"

@app.route('/api/formateurs/stats', methods=['GET'])
@auth_required
def formateurs_stats():
    db = get_db()
    where = []
    params = []
    status = request.args.get('status', 'active')
    if status == 'active':
        where.append('f.is_active = 1')
    elif status == 'inactive':
        where.append('f.is_active = 0')
    etat = request.args.get('etat', '')
    if etat:
        where.append('f.etat = ?')
        params.append(etat)
    search = request.args.get('search', '')
    if search:
        where.append('(f.nom LIKE ? OR f.prenom LIKE ?)')
        params.extend([f'%{search}%', f'%{search}%'])
    competence = request.args.get('competence', '')
    domaine_f = request.args.get('domaine', '')
    join_comp = ''
    if competence or domaine_f:
        join_comp = 'JOIN formateur_competences fc ON fc.formateur_id = f.id'
        if competence:
            where.append('fc.type_competence = ?')
            params.append(competence)
        if domaine_f:
            where.append('fc.domaine = ?')
            params.append(domaine_f)

    if not where:
        where = ['1=1']
    all_where = ' AND '.join(where)

    total = db.execute(f"SELECT COUNT(DISTINCT f.id) as count FROM formateurs f {join_comp} WHERE {all_where}", params).fetchone()['count']
    by_state = [dict(r) for r in db.execute(f"SELECT f.etat, COUNT(DISTINCT f.id) as count FROM formateurs f {join_comp} WHERE {all_where} GROUP BY f.etat ORDER BY count DESC", params).fetchall()]
    by_competence = [dict(r) for r in db.execute(f"SELECT fc2.type_competence, COUNT(DISTINCT fc2.formateur_id) as count FROM formateur_competences fc2 JOIN formateurs f2 ON fc2.formateur_id = f2.id WHERE f2.id IN (SELECT DISTINCT f.id FROM formateurs f {join_comp} WHERE {all_where}) GROUP BY fc2.type_competence ORDER BY count DESC", params).fetchall()]
    db.close()
    return jsonify({'total': total, 'byState': by_state, 'byCompetence': by_competence})

@app.route('/api/formateurs', methods=['GET'])
@auth_required
def list_formateurs():
    db = get_db()
    where = []
    params = []
    status = request.args.get('status', 'active')
    if status == 'active':
        where.append('f.is_active = 1')
    elif status == 'inactive':
        where.append('f.is_active = 0')
    etat = request.args.get('etat', '')
    if etat:
        where.append('f.etat = ?')
        params.append(etat)
    competence = request.args.get('competence', '')
    domaine = request.args.get('domaine', '')
    search = request.args.get('search', '')
    if search:
        where.append("(f.nom LIKE ? OR f.prenom LIKE ?)")
        params.extend([f'%{search}%', f'%{search}%'])
    where_clause = ' AND '.join(where) if where else '1=1'
    page = int(request.args.get('page', 1))
    limit = int(request.args.get('limit', 50))
    offset = (page - 1) * limit
    if competence or domaine:
        join = "INNER JOIN formateur_competences fc ON fc.formateur_id = f.id"
        if competence:
            where_clause += " AND fc.type_competence = ?"
            params.append(competence)
        if domaine:
            where_clause += " AND fc.domaine = ?"
            params.append(domaine)
    else:
        join = ""
    count_sql = f"SELECT COUNT(DISTINCT f.id) as c FROM formateurs f {join} WHERE {where_clause}"
    total = db.execute(count_sql, params).fetchone()['c']
    sql = f"SELECT DISTINCT f.* FROM formateurs f {join} WHERE {where_clause} ORDER BY f.etat, f.nom LIMIT ? OFFSET ?"
    formateurs_rows = db.execute(sql, params + [limit, offset]).fetchall()
    # Batch fetch competences + formations (évite N+1)
    ids = [fr['id'] for fr in formateurs_rows]
    comps_by, forms_by = {}, {}
    if ids:
        placeholders = ','.join(['?'] * len(ids))
        for c in db.execute(f"SELECT * FROM formateur_competences WHERE formateur_id IN ({placeholders})", ids).fetchall():
            comps_by.setdefault(c['formateur_id'], []).append(dict(c))
        for fr_row in db.execute(f"SELECT * FROM formateur_formations WHERE formateur_id IN ({placeholders})", ids).fetchall():
            forms_by.setdefault(fr_row['formateur_id'], []).append(dict(fr_row))
    result = []
    for fr in formateurs_rows:
        d = dict(fr)
        d['competences'] = comps_by.get(fr['id'], [])
        d['formations'] = forms_by.get(fr['id'], [])
        result.append(d)
    db.close()
    return jsonify({'formateurs': result, 'total': total, 'totalPages': max(1, (total + limit - 1) // limit)})

@app.route('/api/formateurs/<int:id>', methods=['GET'])
@auth_required
def get_formateur(id):
    db = get_db()
    f = db.execute("SELECT * FROM formateurs WHERE id = ?", (id,)).fetchone()
    if not f:
        db.close()
        return jsonify({'error': 'Non trouvé'}), 404
    d = dict(f)
    d['competences'] = [dict(c) for c in db.execute("SELECT * FROM formateur_competences WHERE formateur_id = ?", (id,)).fetchall()]
    d['formations'] = [dict(fr) for fr in db.execute("SELECT * FROM formateur_formations WHERE formateur_id = ?", (id,)).fetchall()]
    db.close()
    return jsonify(d)

@app.route('/api/formateurs', methods=['POST'])
@auth_required
def add_formateur():
    if request.user['role'] not in ('National 1', 'Régional', 'Administrateur'):
        return jsonify({'error': 'Non autorisé'}), 403
    db = get_db()
    try:
        nom = request.form.get('nom', '').strip()
        prenom = request.form.get('prenom', '').strip()
        etat = request.form.get('etat', '').strip()
        email = request.form.get('email', '').strip()
        telephone = request.form.get('telephone', '').strip()
        is_inspecteur = 1 if request.form.get('is_inspecteur') == '1' else 0
        if not nom or not prenom or not etat:
            return jsonify({'error': 'Nom, prénom et état requis'}), 400
        ref = generate_formateur_reference(db)
        cv_path = None
        if 'cv' in request.files:
            file = request.files['cv']
            if file.filename:
                ext = file.filename.rsplit('.', 1)[-1] if '.' in file.filename else 'pdf'
                cv_path = f"cv-frm-{secrets.token_hex(8)}.{ext}"
                file.save(os.path.join(UPLOAD_DIR, cv_path))
        db.execute("INSERT INTO formateurs (reference, nom, prenom, etat, email, telephone, cv_path, is_inspecteur) VALUES (?, ?, ?, ?, ?, ?, ?, ?)",
                   (ref, nom, prenom, etat, email, telephone, cv_path, is_inspecteur))
        frm_id = db.lastrowid
        competences_json = request.form.get('competences', '[]')
        for c in json.loads(competences_json):
            if c.get('type_competence'):
                db.execute("INSERT INTO formateur_competences (formateur_id, type_competence, domaine) VALUES (?, ?, ?)",
                           (frm_id, c['type_competence'], c.get('domaine') or None))
        formations_json = request.form.get('formations', '[]')
        for f_item in json.loads(formations_json):
            if f_item.get('description'):
                db.execute("INSERT INTO formateur_formations (formateur_id, type, description) VALUES (?, ?, ?)",
                           (frm_id, f_item.get('type', 'delivree'), f_item['description']))
        log_activity(db, request.user['id'], 'ADD_FORMATEUR', f"Ajout formateur: {ref}")
        db.commit()
        return jsonify({'message': 'Formateur ajouté avec succès', 'id': frm_id, 'reference': ref})
    except Exception as e:
        db.rollback()
        return jsonify({'error': str(e)}), 500
    finally:
        db.close()

@app.route('/api/formateurs/<int:id>', methods=['PUT'])
@auth_required
def update_formateur(id):
    if request.user['role'] not in ('National 1', 'Régional', 'Administrateur'):
        return jsonify({'error': 'Non autorisé'}), 403
    db = get_db()
    try:
        f = db.execute("SELECT * FROM formateurs WHERE id = ?", (id,)).fetchone()
        if not f:
            return jsonify({'error': 'Non trouvé'}), 404
        nom = request.form.get('nom', f['nom']).strip()
        prenom = request.form.get('prenom', f['prenom']).strip()
        etat = request.form.get('etat', f['etat']).strip()
        email = request.form.get('email', f['email'] or '').strip()
        telephone = request.form.get('telephone', f['telephone'] or '').strip()
        is_inspecteur = 1 if request.form.get('is_inspecteur') == '1' else 0
        cv_path = f['cv_path']
        if 'cv' in request.files:
            file = request.files['cv']
            if file.filename:
                ext = file.filename.rsplit('.', 1)[-1] if '.' in file.filename else 'pdf'
                cv_path = f"cv-frm-{secrets.token_hex(8)}.{ext}"
                file.save(os.path.join(UPLOAD_DIR, cv_path))
        db.execute("UPDATE formateurs SET nom=?, prenom=?, etat=?, email=?, telephone=?, cv_path=?, is_inspecteur=?, updated_at=datetime('now') WHERE id=?",
                   (nom, prenom, etat, email, telephone, cv_path, is_inspecteur, id))
        competences_json = request.form.get('competences')
        if competences_json:
            db.execute("DELETE FROM formateur_competences WHERE formateur_id = ?", (id,))
            for c in json.loads(competences_json):
                if c.get('type_competence'):
                    db.execute("INSERT INTO formateur_competences (formateur_id, type_competence, domaine) VALUES (?, ?, ?)",
                               (id, c['type_competence'], c.get('domaine') or None))
        formations_json = request.form.get('formations')
        if formations_json:
            db.execute("DELETE FROM formateur_formations WHERE formateur_id = ?", (id,))
            for f_item in json.loads(formations_json):
                if f_item.get('description'):
                    db.execute("INSERT INTO formateur_formations (formateur_id, type, description) VALUES (?, ?, ?)",
                               (id, f_item.get('type', 'delivree'), f_item['description']))
        log_activity(db, request.user['id'], 'UPDATE_FORMATEUR', f"Modification formateur: {f['reference']}")
        db.commit()
        return jsonify({'message': 'Formateur modifié avec succès'})
    except Exception as e:
        db.rollback()
        return jsonify({'error': str(e)}), 500
    finally:
        db.close()

@app.route('/api/formateurs/<int:id>/deactivate', methods=['PUT'])
@auth_required
def deactivate_formateur(id):
    if request.user['role'] != 'Administrateur':
        return jsonify({'error': 'Non autorisé'}), 403
    db = get_db()
    db.execute("UPDATE formateurs SET is_active = 0, updated_at = datetime('now') WHERE id = ?", (id,))
    log_activity(db, request.user['id'], 'DEACTIVATE_FORMATEUR', f"Désactivation formateur #{id}")
    db.commit()
    db.close()
    return jsonify({'message': 'Formateur désactivé'})

@app.route('/api/formateurs/<int:id>/activate', methods=['PUT'])
@auth_required
def activate_formateur(id):
    if request.user['role'] != 'Administrateur':
        return jsonify({'error': 'Non autorisé'}), 403
    db = get_db()
    db.execute("UPDATE formateurs SET is_active = 1, updated_at = datetime('now') WHERE id = ?", (id,))
    log_activity(db, request.user['id'], 'ACTIVATE_FORMATEUR', f"Réactivation formateur #{id}")
    db.commit()
    db.close()
    return jsonify({'message': 'Formateur réactivé'})

@app.route('/api/formateurs/bulk-delete', methods=['POST'])
@auth_required
def bulk_delete_formateurs():
    if request.user['role'] != 'Administrateur':
        return jsonify({'error': 'Non autorisé'}), 403
    ids = request.json.get('ids', [])
    if not ids:
        return jsonify({'error': 'Aucun formateur sélectionné'}), 400
    db = get_db()
    placeholders = ','.join(['?'] * len(ids))
    db.execute(f"DELETE FROM formateur_competences WHERE formateur_id IN ({placeholders})", ids)
    db.execute(f"DELETE FROM formateur_formations WHERE formateur_id IN ({placeholders})", ids)
    db.execute(f"DELETE FROM formateurs WHERE id IN ({placeholders})", ids)
    log_activity(db, request.user['id'], 'BULK_DELETE_FORMATEURS', f"Suppression de {len(ids)} formateur(s)")
    db.commit()
    db.close()
    return jsonify({'message': f'{len(ids)} formateur(s) supprimé(s) avec succès'})

@app.route('/api/formateurs/<int:id>/email', methods=['POST'])
@auth_required
def send_formateur_email(id):
    if request.user['role'] not in ('Régional', 'Administrateur'):
        return jsonify({'error': 'Non autorisé'}), 403
    db = get_db()
    f = db.execute("SELECT * FROM formateurs WHERE id = ?", (id,)).fetchone()
    if not f or not f['email']:
        db.close()
        return jsonify({'error': "Pas d'email"}), 400
    data = request.json
    subject = data.get('subject', '')
    body = data.get('body', '')
    log_activity(db, request.user['id'], 'SEND_EMAIL_FORMATEUR', f"Email à {f['email']}: {subject}")
    db.commit()
    db.close()
    mailto = f"mailto:{f['email']}?subject={urllib.parse.quote(subject)}&body={urllib.parse.quote(body)}"
    return jsonify({'message': f"Ouverture du client email pour {f['email']}", 'mailto': mailto})

@app.route('/api/formateurs/export/excel', methods=['GET'])
@auth_required
def export_formateurs_excel():
    db = get_db()
    rows = db.execute("SELECT f.*, GROUP_CONCAT(DISTINCT fc.type_competence || ' - ' || fc.domaine) as competences_str FROM formateurs f LEFT JOIN formateur_competences fc ON fc.formateur_id = f.id WHERE f.is_active = 1 GROUP BY f.id ORDER BY f.etat, f.nom").fetchall()
    db.close()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Formateurs'
    header_font = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='1A365D', end_color='1A365D', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(left=Side(style='thin', color='D0D0D0'), right=Side(style='thin', color='D0D0D0'), top=Side(style='thin', color='D0D0D0'), bottom=Side(style='thin', color='D0D0D0'))
    ws.merge_cells('A1:H1')
    ws['A1'].value = 'UEMOA - Base de données des Formateurs'
    ws['A1'].font = Font(name='Calibri', bold=True, size=14, color='1A365D')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 35
    ws.merge_cells('A2:H2')
    ws['A2'].value = 'Exporté le ' + datetime.now().strftime(_DT_FMT)
    ws['A2'].font = Font(name='Calibri', italic=True, size=10, color='666666')
    ws['A2'].alignment = Alignment(horizontal='center')
    ws.row_dimensions[3].height = 8
    headers = ['Référence', 'Nom', 'Prénom', 'État', 'Email', 'Téléphone', 'Compétences', 'Inspecteur']
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    alt_fill = PatternFill(start_color='F7FAFC', end_color='F7FAFC', fill_type='solid')
    for row_idx, r in enumerate(rows, 5):
        data = [r['reference'], r['nom'], r['prenom'], r['etat'], r['email'] or '', r['telephone'] or '', r['competences_str'] or '', 'Oui' if r['is_inspecteur'] else 'Non']
        for col_idx, value in enumerate(data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = Font(name='Calibri', size=10)
            cell.border = thin_border
            cell.alignment = Alignment(vertical='center', wrap_text=True)
            if (row_idx - 5) % 2 == 1:
                cell.fill = alt_fill
    col_widths = [16, 18, 18, 16, 28, 18, 40, 12]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.freeze_panes = 'A5'
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True, download_name='formateurs.xlsx')

@app.route('/api/formateurs/export/csv', methods=['GET'])
@auth_required
def export_formateurs_csv():
    db = get_db()
    rows = db.execute("SELECT f.*, GROUP_CONCAT(DISTINCT fc.type_competence || COALESCE(' - ' || fc.domaine, '')) as competences_str FROM formateurs f LEFT JOIN formateur_competences fc ON fc.formateur_id = f.id WHERE f.is_active = 1 GROUP BY f.id ORDER BY f.etat, f.nom").fetchall()
    db.close()
    output = io.StringIO()
    output.write('\ufeff')
    writer = csv.writer(output, delimiter=';')
    writer.writerow(['Référence', 'Nom', 'Prénom', 'État', 'Email', 'Téléphone', 'Compétences', 'Inspecteur'])
    for r in rows:
        writer.writerow([r['reference'], r['nom'], r['prenom'], r['etat'], r['email'] or '', r['telephone'] or '', r['competences_str'] or '', 'Oui' if r['is_inspecteur'] else 'Non'])
    output.seek(0)
    return send_file(io.BytesIO(output.getvalue().encode('utf-8-sig')),
        mimetype='text/csv', as_attachment=True, download_name='formateurs.csv')

@app.route('/api/formateurs/export/pdf', methods=['GET'])
@auth_required
def export_formateurs_pdf():
    db = get_db()
    rows = db.execute("SELECT f.*, GROUP_CONCAT(DISTINCT fc.type_competence || COALESCE(' - ' || fc.domaine, '')) as competences_str FROM formateurs f LEFT JOIN formateur_competences fc ON fc.formateur_id = f.id WHERE f.is_active = 1 GROUP BY f.id ORDER BY f.etat, f.nom").fetchall()
    db.close()

    def clean(text):
        if not text:
            return ''
        text = str(text)
        replacements = {
            '\u2019': "'", '\u2018': "'", '\u201c': '"', '\u201d': '"',
            '\u2013': '-', '\u2014': '-', '\u2026': '...', '\u00a0': ' ',
            '\u2022': '-', '\u200b': '', '\u200e': '', '\u200f': '',
            '\u2010': '-', '\u2011': '-', '\u2012': '-', '\u2015': '-',
            '\u2032': "'", '\u2033': '"', '\u02bc': "'", '\u02bb': "'",
            '\ufeff': '', '\u200c': '', '\u200d': '', '\u202a': '', '\u202c': '',
        }
        for old_c, new_c in replacements.items():
            text = text.replace(old_c, new_c)
        try:
            text.encode('latin-1')
        except UnicodeEncodeError:
            text = text.encode('latin-1', 'replace').decode('latin-1')
        return text

    FONT_DIR = 'C:/Windows/Fonts' if os.path.exists('C:/Windows/Fonts') else '/usr/share/fonts/truetype/dejavu'

    class FormateursPDF(FPDF):
        def __init__(self, *args, **kwargs):
            super().__init__(*args, **kwargs)
            if os.path.exists(os.path.join(FONT_DIR, 'arial.ttf')):
                self.add_font('ArialUni', '', os.path.join(FONT_DIR, 'arial.ttf'))
                self.add_font('ArialUni', 'B', os.path.join(FONT_DIR, 'arialbd.ttf'))
                self.add_font('ArialUni', 'I', os.path.join(FONT_DIR, 'ariali.ttf'))
            elif os.path.exists(os.path.join(FONT_DIR, 'DejaVuSans.ttf')):
                self.add_font('ArialUni', '', os.path.join(FONT_DIR, 'DejaVuSans.ttf'))
                self.add_font('ArialUni', 'B', os.path.join(FONT_DIR, 'DejaVuSans-Bold.ttf'))
                self.add_font('ArialUni', 'I', os.path.join(FONT_DIR, 'DejaVuSans-Oblique.ttf'))

        def header(self):
            self.set_font('ArialUni', 'B', 14)
            self.set_text_color(26, 54, 93)
            self.cell(0, 8, clean('UEMOA - Union Economique et Monetaire Ouest Africaine'), new_x="LMARGIN", new_y="NEXT", align='C')
            self.set_font('ArialUni', '', 10)
            self.set_text_color(43, 87, 151)
            self.cell(0, 6, clean('Base de donnees des Formateurs'), new_x="LMARGIN", new_y="NEXT", align='C')
            self.set_font('ArialUni', 'I', 8)
            self.set_text_color(113, 128, 150)
            self.cell(0, 5, clean('Exporté le ' + datetime.now().strftime(_DT_FMT) + f' | Total: {len(rows)} enregistrement(s)'), new_x="LMARGIN", new_y="NEXT", align='C')
            self.set_draw_color(26, 54, 93)
            self.set_line_width(0.5)
            self.line(self.l_margin, self.get_y() + 2, self.w - self.r_margin, self.get_y() + 2)
            self.ln(5)

        def footer(self):
            self.set_y(-12)
            self.set_font('ArialUni', 'I', 7)
            self.set_text_color(113, 128, 150)
            self.cell(0, 10, clean(f"UEMOA - URSAC | Page {self.page_no()}/{{nb}}"), align='C')

    pdf = FormateursPDF(orientation='L', format='A4')
    pdf.alias_nb_pages()
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.add_page()

    cols = [('Ref.', 24), ('Nom', 32), ('Prenom', 32), ('Etat', 24), ('Email', 42), ('Tel.', 24), ('Competences', 70), ('Inspecteur', 18)]

    pdf.set_font('ArialUni', 'B', 7)
    pdf.set_fill_color(26, 54, 93)
    pdf.set_text_color(255, 255, 255)
    for name, w in cols:
        pdf.cell(w, 7, name, border=1, fill=True, align='C')
    pdf.ln()

    pdf.set_font('ArialUni', '', 6.5)
    fill = False
    for r in rows:
        if pdf.get_y() > pdf.h - 20:
            pdf.add_page()
            pdf.set_font('ArialUni', 'B', 7)
            pdf.set_fill_color(26, 54, 93)
            pdf.set_text_color(255, 255, 255)
            for name, w in cols:
                pdf.cell(w, 7, name, border=1, fill=True, align='C')
            pdf.ln()
            pdf.set_font('ArialUni', '', 6.5)
            fill = False

        if fill:
            pdf.set_fill_color(247, 250, 252)
        else:
            pdf.set_fill_color(255, 255, 255)

        pdf.set_text_color(26, 32, 44)
        data = [
            clean(r['reference']), clean(r['nom']), clean(r['prenom']),
            clean(r['etat']), clean(r['email']), clean(r['telephone']),
            clean((r['competences_str'] or '')[:55]),
            'Oui' if r['is_inspecteur'] else 'Non',
        ]
        for i, (name, w) in enumerate(cols):
            pdf.cell(w, 6, data[i], border='TB', fill=True)
        pdf.ln()
        fill = not fill

    output = io.BytesIO()
    pdf.output(output)
    output.seek(0)
    return send_file(output, mimetype='application/pdf', as_attachment=True, download_name='formateurs.pdf')

@app.route('/api/formateurs/import', methods=['POST'])
@auth_required
def import_formateurs():
    if request.user['role'] != 'Administrateur':
        return jsonify({'error': 'Non autorisé'}), 403
    file = request.files.get('file')
    if not file:
        return jsonify({'error': 'Fichier requis'}), 400
    try:
        wb = openpyxl.load_workbook(file)
        ws = wb.active
        db = get_db()
        count = 0
        for row in ws.iter_rows(min_row=5, values_only=True):
            if not row or not row[1]:
                continue
            nom = str(row[1] or '').strip()
            prenom = str(row[2] or '').strip()
            etat = str(row[3] or '').strip()
            email = str(row[4] or '').strip()
            telephone = str(row[5] or '').strip()
            is_insp_raw = str(row[6] if len(row) > 6 else '').strip().lower()
            type_competence = str(row[7] if len(row) > 7 else '').strip()
            comp_domaine = str(row[8] if len(row) > 8 else '').strip()
            f_delivrees = str(row[9] if len(row) > 9 else '').strip()
            f_developpees = str(row[10] if len(row) > 10 else '').strip()
            if not nom or not etat:
                continue
            is_insp = 1 if is_insp_raw in ('oui', '1', 'yes', 'true') else 0
            existing = db.execute("SELECT id FROM formateurs WHERE nom = ? AND prenom = ? AND etat = ?", (nom, prenom, etat)).fetchone()
            if existing:
                f_id = existing['id']
                db.execute("UPDATE formateurs SET email = ?, telephone = ?, is_inspecteur = ?, updated_at = datetime('now') WHERE id = ?",
                           (email, telephone, is_insp, f_id))
            else:
                ref = generate_formateur_reference(db)
                db.execute("INSERT INTO formateurs (reference, nom, prenom, etat, email, telephone, is_inspecteur) VALUES (?, ?, ?, ?, ?, ?, ?)",
                           (ref, nom, prenom, etat, email, telephone, is_insp))
                f_id = db.lastrowid
                count += 1
            if type_competence:
                db.execute("INSERT INTO formateur_competences (formateur_id, type_competence, domaine) VALUES (?, ?, ?)",
                           (f_id, type_competence, comp_domaine or None))
            for desc in [s.strip() for s in f_delivrees.split(';') if s.strip()]:
                db.execute("INSERT INTO formateur_formations (formateur_id, type, description) VALUES (?, 'delivree', ?)", (f_id, desc))
            for desc in [s.strip() for s in f_developpees.split(';') if s.strip()]:
                db.execute("INSERT INTO formateur_formations (formateur_id, type, description) VALUES (?, 'developpee', ?)", (f_id, desc))
        log_activity(db, request.user['id'], 'IMPORT_FORMATEURS', f"Import Excel: {count} nouveaux formateurs")
        db.commit()
        db.close()
        return jsonify({'message': f'{count} formateur(s) importé(s) avec succès'})
    except Exception as e:
        return jsonify({'error': f"Erreur d'import: {str(e)}"}), 400


# Initialize database on import (for gunicorn) and on direct run
import time as _time

def _init_db_with_retry(retries=10, delay=4):
    for attempt in range(1, retries + 1):
        try:
            init_db()
            print(f"Base de données initialisée (tentative {attempt})")
            return
        except Exception as e:
            print(f"[init_db] Tentative {attempt}/{retries} échouée : {e}")
            if attempt < retries:
                _time.sleep(delay)
            else:
                print("[init_db] Impossible d'initialiser la base après plusieurs tentatives.")
                raise

_init_db_with_retry()

if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5005))
    print(f"Démarrage du serveur sur http://localhost:{port}")
    app.run(host='0.0.0.0', port=port, debug=False)
